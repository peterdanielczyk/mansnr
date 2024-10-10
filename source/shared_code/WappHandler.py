try:
    import datetime
    import importlib
    import threading
    import traceback

    #import shared_code.Services
    from shared_code.AzureHelper import CONT, XEDB, dumps, hasattr, loads
except:
    from AzureHelper import loads,dumps,CONT,XEDB,hasattr
    import datetime,threading,importlib,traceback,traceback

import base64
import datetime
import json
import logging
import os
import re
import time
from datetime import timedelta
from importlib import reload
from io import BytesIO
from typing import Any, List, Optional

import holidays
import psycopg2
import pytz
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo

#import shared_code.coolfish_statics

BLMAP={
    "Baden-Württemberg":"BW",
    "Bayern":"BY",
    "Berlin":"BE",
    "Brandenburg":"BR",
    "Bremen":"HB",
    "Hamburg":"HH",
    "Hessen":"HE",
    "Mecklenburg-Vorpommern":"MV",
    "Niedersachsen":"NI",
    "Nordrhein-Westfalen":"NW",
    "Rheinland-Pfalz":"RP",
    "Saarland":"SL",
    "Sachsen":"SN",
    "Sachsen-Anhalt":"ST",
    "Schleswig-Holstein":"SH",
    "Thüringen":"TH"
}

BLKLAR=BLMAP.keys()

BLMAPR = {v: k for k, v in BLMAP.items()}

def add_years(d, years):
    """Return a date that's `years` years after the date (or datetime)
    object `d`. Return the same calendar date (month and day) in the
    destination year, if it exists, otherwise use the following day
    (thus changing February 29 to March 1).

    """
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        return d + (datetime.date(d.year + years, 1, 1) - datetime.date(d.year, 1, 1))

def num2month(nm: int) -> (str):
    """
    Converts a month number to its corresponding name in German.

    Parameters:
    - nm (int): The month number (1 for January, 2 for February, etc.).

    Returns:
    - str: The German name of the month.
    """
    months = ["Januar", "Februar", "März", "April", "Mai", "Juni", 
              "Juli", "August", "September", "Oktober", "November", "Dezember"]
    return months[nm - 1]

def num2monthEng(nm: int) -> (str):
    """
    Converts a month number to its corresponding name in English.

    Parameters:
    - nm (int): The month number (1 for January, 2 for February, etc.).

    Returns:
    - str: The English abbreviation of the month.
    """
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", 
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dez"]
    return months[nm - 1]

def month2num(month: str) -> (int):
    """
    Converts a month's abbreviated English name to its corresponding number.

    Parameters:
    - month (str): The abbreviated English name of the month (e.g., "Jan" for January).

    Returns:
    - int: The month number (1 for January, 2 for February, etc.).
    """
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", 
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return months.index(month) + 1

def calPermissions(selfqnr, getp=1):
    """
    Calculates permissions related to a specific Q-Number based on the permissions stored in a cache.

    This method iterates over cached permission objects to determine which permissions are applicable
    to a given Q-Number, either as a source or a target of the permission, depending on
    the mode specified by the getp parameter.

    Parameters:
    - selfqnr: The Q-Number for which permissions are being calculated.
    - getp (optional): A flag determining the mode of permission calculation. If set to 1 (default),
                       the method looks for permissions where selfqnr is the target. If set to 0,
                       it looks for permissions where selfqnr is the source.

    Returns:
    - A dictionary of permissions related to the specified Q-Number, keyed by the relevant Q-Number in the
      permission relation (source or target, based on getp).
    """
    
    # Initialize an empty dictionary to hold the permissions
    spermissions = {}
    
    # Check if the 'cal_permission' category exists in the object cache
    if "cal_permission" in getOCACHE():
        # Iterate over all permissions in the cache
        for p in getOCACHE()["cal_permission"].values():
            # Skip the loop iteration if the permission object doesn't have the 'permission' attribute
            if not hasattr(p, "permission"):
                continue
            
            # Load the permission data (expecting it to be a JSON string)
            perm = loads(p.permission)
            
            # If in 'target' mode and the current permission's target matches selfqnr, add it to the dictionary
            if getp:
                if perm[2] == selfqnr:
                    spermissions[perm[1]] = p
            # If in 'source' mode and the current permission's source matches selfqnr, add it to the dictionary
            else:
                if perm[1] == selfqnr:
                    spermissions[perm[2]] = p
    
    # Return the dictionary of permissions
    return spermissions

def getMAbyQnr(qnr):
    for mtype in ["mitarbeiter","nomiko","planzak"]:
        ma=getOCACHE()["%s_qnr"%mtype].get(qnr)
        if ma: return ma
    return None


def getCurrentFK(ma_or_qnr, isma=False):
    """
    Retrieve the current supervisor (Führungskraft) for a given employee.

    Args:
        ma_or_qnr: Either a Mitarbeiter object or a Q-Nummer (employee ID).
        isma (bool): Indicates whether the first argument is a Mitarbeiter object.

    Returns:
        The current supervisor (Führungskraft) as a Mitarbeiter object, or None if not found.
    """
    # Get the current date as a string in the format 'YYYY-MM-DD'
    dtnowstr = datetime.date.today().strftime("%Y-%m-%d")

    # Determine the Mitarbeiter object from Q-Nummer if not provided directly
    ma = getMAbyQnr(ma_or_qnr) if not isma and not isCont(ma_or_qnr) else ma_or_qnr

    # Retrieve the team associated with the employee
    vteam = getS(ma, "team")

    # If no team is found, retrieve the contract and team information based on the current date
    if not vteam:
        v = getVertragByDate(ma.q_nummer, dtnowstr)
        if not v or not getS(v, "team"):
            return None
        vteam = v.team

    # Get the organizational unit for the team as of the current date
    team = getOrgByDate(ma, "team", vteam, dtnowstr)

    # If no team or Q-Nummer is found, return None
    if not team or not team.q_nummer:
        return None
    
    # If the team leader's Q-Nummer is the same as the employee's, check the department (Abteilung)
    if team.q_nummer == ma.q_nummer:
        abtlg = getOrgByDate(ma, "abtlg", v.abteilung, dtnowstr)
        # If the department leader's Q-Nummer is also the same, return None
        if abtlg.q_nummer == ma.q_nummer:
            return None

def getAbwByQNR(qnr, start, end):
    """
    Retrieve absence data for a specific employee within a given date range.

    Args:
        qnr: The Q-Nummer (employee ID) for which to retrieve absence data.
        start: The start date of the period to check (as a datetime object).
        end: The end date of the period to check (as a datetime object).

    Returns:
        A dictionary where the keys are absence types and the values are the total days of absence for each type within the specified date range.
    """
    # Check if absence data is already cached; if not, load it into the cache
    if not getOCACHE().get("abwesenheit_qnr"):
        getOCACHE()["abwesenheit_qnr"] = dictByPNumList("abwesenheit")
    
    # Retrieve the list of absences for the specified employee (Q-Nummer)
    al = getOCACHE()["abwesenheit_qnr"].get(qnr)
    if not al:
        return {}
    
    ret = {}
    # Sort absences by type, in reverse order
    al.sort(key=lambda x: x.art, reverse=True)
    
    for a in al:
        # Extend the absence period if necessary
        extendAbw(a)
        
        # Get the start and end dates of the absence
        st1 = getD(a, "erster_abwesen_heitstag")
        en1 = getD(a, "letzter_abwesen_heitstag")
        
        # Skip absences without valid start or end dates
        if not en1 or not st1:
            continue
        
        # Skip absences outside the specified date range
        if st1 > end or en1 < start:
            continue
        
        # Determine the overlapping period of the absence within the specified date range
        st2 = max(start, st1)
        en2 = min(end, en1)
        
        # Calculate the number of days of absence within the overlapping period
        r = getOrAdd(ret, a.art, 0)
        r += (en2 - st2).days + 1
        
        # Adjust for part-time during parental leave
        if a.art == "Elternzeit":
            r -= ret.get("Teilzeit in Elternzeit", 0)
        
        # Store the total days of absence for the absence type
        ret[a.art] = r
    
    return ret

def daysAbsenceCMonth(iparam: Any, alist: List[Any]) -> int:
    """
    Calculates the total number of absence days for each 'abwesenheit' entry within a given month.
    
    The function iterates over a list of absence records and calculates the total number of days absent
    in the context of a specific month defined by 'iparam'. It handles edge cases where the absence period
    extends beyond the start or end of the month.
    
    Args:
        iparam: A parameter object containing attributes like 'ad_start', 'ad_end', and 'eomday' 
                which respectively represent the start of the period, the end of the period, 
                and the end of month day count.
        alist: A list of 'abwesenheit' entries where each entry is an object containing 
               'erster_abwesen_heitstag' (start of absence) and 'letzter_abwesen_heitstag' (end of absence).
    
    Returns:
        The total number of absence days in the month for the first entry that meets the criteria.
        If no entries meet the criteria, returns 0. This function currently returns after evaluating
        the first valid entry in the list; this behavior may need adjustment based on intended use.
        
    Note:
        This function returns after checking the first valid 'abwesenheit' entry. If the intention is to
        sum or evaluate all entries, the logic would need to be adjusted accordingly.
    """
    for a in alist:
        st = getD(a, "erster_abwesen_heitstag")  # Start date of absence
        en = getD(a, "letzter_abwesen_heitstag")  # End date of absence
        
        if st and en:
            # Calculate the number of days past the end of the month from the absence end date
            d2 = (en - iparam.ad_end).days
            # If the absence ends before the month starts, return 0
            if d2 < -iparam.eomday:
                return 0
            # If the absence end date is after the month end, adjust d2 to 0
            if d2 > 0:
                d2 = 0
            
            # Calculate the number of days before the start of the month from the absence start date
            d1 = (iparam.ad_start - st).days
            # If the absence starts after the month ends, return 0
            if d1 < -iparam.eomday:
                return 0
            # If the absence start date is before the month start, adjust d1 to 0
            if d1 > 0:
                d1 = 0
            
            # Return the total days of absence within the month
            return iparam.eomday + d1 + d2
    return 0  # Return 0 if no valid 'abwesenheit' entries are found


def getDataFromDatabaseRaw(sql,verbose=0):
    '''* This function returns data based on a raw sql statement.
    * It uses much less ``hidden_magic`` than most other database calls in the framework.'''    
    if verbose:print(sql)
    cursor = XEDB.DBConn.cursor()
    cursor.execute(sql)
    result = cursor.fetchall() if cursor.description else None    
    XEDB.DBConn.commit()
    return result


if 0:

    def hasattr(o,n):
        if not o:return False
        return n in o.__dict__

    class CONT:
        def __init__(self) -> None:
            self.__persist=0

        def __getattr__(self, name):
            if name in self.__dict__:
                # Default behaviour
                return self.__getattribute__(self, name)
            else:
                return None
        def persistOn(self):
            self.__persist=1
        def persistOff(self):
            self.__persist=0
        def __setattr__(self, name,val):
            if self.__dict__ and self.__persist:
                iparam=CONT()
                iparam.user="CACHE_ADD"
                setObjectAttribute(iparam,self,name,val,ifdiff=1)
            # Default behaviour
            object.__setattr__(self, name, val)

    try:
        XEDB
    except:
        XEDB=CONT()

    def getIParam(req):
        req_body = req.get_body()
        return loads(req_body)

    def obj2dict(pp):
        if isinstance(pp,tuple):
            olist=[]
            for p in pp:
                olist.append(obj2dict(p))
            return tuple(olist)
        if isinstance(pp,list):#type(pp) is types.ListType:
            olist=[]
            for p in pp:
                olist.append(obj2dict(p))
            return olist

        if isinstance(pp,dict):#type(pp) is types.DictionaryType:
            odict={}
            for k,val in pp.items():
                odict[k]=obj2dict(val)
            return odict
        if isinstance(pp,CONT): #inspect.isclass(pp): #:#type(pp) is types.InstanceType:
            odict={}
            for m in pp.__dict__:
                val=getattr(pp,m)
                odict[m]=obj2dict(val)
            return odict
        return pp


    def dict2obj(pp):
        if isinstance(pp,list):# type(pp) is types.ListType:
            olist=[]
            for p in pp:
                olist.append(dict2obj(p))
            return olist
        if isinstance(pp,dict):
            oinst=CONT()
            for (k,v) in pp.items():
                val=dict2obj(v)
                setattr(oinst,k,val)
            return oinst
        return pp

    def dumps(pp):
        '''* Returns a ``json`` string.
        * Expects a ``CONT object``, because it first callse ``ob2dict`` method on ``pp``.'''
        return json.dumps(obj2dict(pp))
    
    def loads(pp):
        '''* Returns a ``CONT object`.
        * Expects a valid ``json``, because it first callse ``dict2obj`` method on ``pp``.'''
        return dict2obj(json.loads(pp))

import zlib


def zipStringBase64(ss):
    ss=zlib.compress(ss.encode())
    ss=base64.b64encode(ss).decode()
    return ss

def unzipStringBase64(ss):
    ss=base64.b64decode(ss.encode())
    ss=zlib.decompress(ss).decode()
    return ss

def object2bas64Zip(fo):
    dummy=dumps(fo)
    return zipStringBase64(dummy)

def object2bas64(fo):
    dummy=dumps(fo)
    dummy=dummy.encode()
    return base64.b64encode(dummy).decode()

def bas642object(val_string):
    return loads(base64.b64decode(val_string))

def poolActive(pool: str, pools: dict):
    '''Determines if a 360Feedback Pool is active now.
    '''

    pool_active=bool(1)

    if pool=="Allgemein":
        return "ok"

    now = datetime.datetime.now()
    activation_period_duration=21

    if pool not in pools:
            pool_active=bool(0)
            return pool_active

    if pools[pool]:
        pool_start=datetime.datetime.strptime(pools[pool], '%Y-%m-%d')
        pool_end=pool_start+datetime.timedelta(days=activation_period_duration,hours=23,minutes=59)

    pool_active="ok"
    if pool_start > now:
        pool_active="notyet"
    if pool_end < now:
        pool_active="ok"

    return pool_active

def getPoolStartDate(pool: str, pools: dict):
    '''Returns the start date for a pool.
    '''
    pool_startdate=None

    if pool=="Allgemein":
        pool_startdate=None
        return pool

    if pool not in pools:
            pool_startdate=None
            return pool_startdate

    if pools[pool]:
        pool_startdate=datetime.datetime.strptime(pools[pool], '%Y-%m-%d')

    return pool_startdate

def getPoolEndDate(pool: str, pools: dict):
    '''Returns the end date for a pool.
    '''
    pool_enddate=None
    activation_period_duration=21
    pool_startdate=getPoolStartDate(pool, pools)

    if pool=="Allgemein":
        pool_enddate=None
        return pool_enddate

    if pool not in pools:
        pool_enddate=None
        return pool_enddate

    if pools[pool]:
        pool_enddate=datetime.datetime.strptime(pools[pool], '%Y-%m-%d')
        pool_enddate=pool_startdate+datetime.timedelta(days=activation_period_duration)

    return pool_enddate.strftime(DFMT)

def getPoolNumber(q_nummer):
    '''Gets the 360Grad FB Pool for a specified ``q_nummer``
    '''
    pool=None
    data=getObjects("grad360fb")
    for feedback in data:
        if q_nummer==feedback.q_nummer==feedback.provider:
            pool=feedback.pool
    return str(pool)

def getPoolDates(q_nummer):
        from Services.getGrad360prov import POOLSTARTDATES
        my_pool=getPoolNumber(q_nummer)
        ret=CONT()
        ret.startdate=str(getPoolStartDate(my_pool,POOLSTARTDATES))[0:11]
        ret.enddate=str(getPoolEndDate(my_pool,POOLSTARTDATES))[0:11]
        return ret

def setObjectAttribute(iparam,object,attr,val,ifdiff=0):
    '''Sets an attribute to a specific value for an object.
    - ``object`` needs to be a specific object
    - ``attr`` is the attribute that should be set/changed
    - ``val`` is the desired new value
    - ``ifdiff`` defines that the new value should only be set if it is different from the old one'''
    if not object.oid:return None
    rd=genericSelect("select distinct otype from public.excelse_%s where oid='%s' and status is null"%(XEDB.DOMAIN,object.oid))
    if ifdiff:
        getObject(rd[0].otype,object.oid)
        if str(getattr(object,attr))==str(val):
            return None
    newobject=CONT()
    newobject.oid=object.oid
    setattr(newobject,attr,val)
    ret=setObject(iparam, rd[0].otype, newobject)
    if not hasattr(object,attr) or getattr(object,attr)!=val:
        setattr(object,attr,val)
    return ret


class CFTable:
    def __init__(self) -> None:
        self.row=0

    def setObject(self,o):
        if not self.row:
            self.row=self.startRow
        for attr in o.__dict__:
            col=self.NAMESCOL.get(attr)
            if col:
                sh=self.sheet
                val=getattr(o,attr)
                if isCont(val):val=val.value
                sh.cell(self.row,col+self.startColumn-1).value=val
        self.row+=1

    def clearContent(self):
        for row in range(self.startRow,self.endRow+1):
            for col in range(self.startColumn,self.endColumn+1):
                self.sheet.cell(row,col).value=None

def deleteObject(oid,DMARK="SYSDEL"):
    otype=getOtypeByOID(oid)
    if not otype:
        logging.error("No otype found for OID {}".format(oid))
        return False
    else:
        otype=otype.strip()
        sql="update public.excelse_%s set status='%s' where otype='%s' and oid='%s' "%(XEDB.DOMAIN,DMARK,otype,oid)
        logging.info("Successfully deleted object with OID {} from Database".format(oid))
        print(sql)
        runDBStatement(sql)
        try:
            del getOCACHE()[otype][oid]
            logging.info("Successfully deleted object with OID {} from OCACHE".format(oid))
        except:
            logging.info("Cannot delete object of type %s with OID %s from OCACHE"%(otype,oid))
        return True

from math import ceil

factor_of_font_size_to_width = {
    12: {
        "factor": 0.8,  # width / count of symbols at row
        "height": 16
    },
    11: {
        "factor": 0.7,  # width / count of symbols at row
        "height": 15
    }
}

class MEXCEL:
    def __init__(self,fname=None) -> None:
        if fname:
            self.filename=fname
            self.wb = load_workbook(filename = fname)
            self.tables={}
            self.asheet=self.wb.worksheets[0]

    def setAutoHeight(self,row_number,font_size,cwidth,stretch=1.0):
        font_params = factor_of_font_size_to_width[font_size]

        row = list(self.asheet.rows)[row_number-1]
        height = font_params["height"]

        for cell in row:
            words_count_at_one_row = cwidth / font_params["factor"]
            lines = ceil(len(str(cell.value)) / words_count_at_one_row)
            height = max(height, lines * font_params["height"])
        if height>font_params["height"]:self.asheet.row_dimensions[row_number].height=height*stretch

    def setVal(self,row,col,val):
        if isinstance(val, str):
            val=val.encode("utf-8","ignore")
        self.asheet.cell(row,col).value=val

    def setValByName(self,name,val):
        rc=self.getRowColFromName(name)
        self.setVal(rc.row,rc.col,val)

    def getRowColFromName(self,name):
        ret=CONT()
        my_range = self.wb.defined_names[name]
        dests = my_range.destinations # returns a generator of (worksheet title, cell range) tuples
        cells = []
        for title, coord in dests:
            print()
        cl=coord.split("$")
        ret.col=ColRef2ColNo(cl[1])
        ret.row=int(cl[2])
        return ret

    def setActiveSheet(self,sh):
        dummy=type(sh)
        self.asheet=sh

    def getTable(self,tname=None):
        ct= self.tables.get(tname)
        if ct:
            return ct
        t=None
        breakit=0
        for sh in self.wb.worksheets:
            if breakit:break
            if sh.tables:
                t=list(sh.tables.values())[0]

            for tname_in,t in dict(sh.tables).items():
                if tname==tname_in or not tname:
                    breakit=1
                    ct=self.tables[t.name]=CFTable()
                    ct.table=t
                    ct.sheet=sh
                    break
        if t:
            trefl=t.ref.split(":")
            ct.startRow=coordinate_from_string(trefl[0])[1]+1
            ct.endRow=coordinate_from_string(trefl[1])[1]
            ct.startColumn=ColRef2ColNo(coordinate_from_string(trefl[0])[0])
            ct.endColumn=len(t.tableColumns)+ct.startColumn-1
            ct.NAMESCOL={}
            ct.COLNAMES={}
            ct.xls=self
            i=0
            for cn in t.column_names:
                i+=1
                cn=namesToVars([cn])[0]
                ct.NAMESCOL[cn]=i
                ct.COLNAMES[i]=cn
        
        return ct

    def tableObjects(self,tname=None):
        t=self.getTable(tname)
        for row in range(t.startRow,t.endRow+1):
            obj=CONT()
            for col in range(t.startColumn,t.endColumn+1):
                cellVal=self.wb.active.cell(row,col).value
                setattr(obj,t.COLNAMES[col],cellVal)
            yield obj

    def save(self,fn=None):
        if not fn:fn=self.filename
        self.wb.save(fn)

    def saveAsStream(self,ret):
        virtual_workbook = BytesIO()
        self.wb.save(virtual_workbook)
        octets=virtual_workbook.getvalue()
        ret.filelength=len(octets)
        octets=base64.b64encode(octets)
        ostring=""
        for o in octets:
            ostring+=chr(o)
        ret.stream=ostring


def ColRef2ColNo(cr):
    cr=cr.upper()
    if len(cr)==2:
        return (ord(cr[0])-ord('A')+1)*26 + ord(cr[1])-ord('A')+1
    return ord(cr)-ord('A')+1

def colnum2Letter(cn):
    numAs=int((cn-1)/26)
    ret=""
    if numAs:
        ret+=chr(64+numAs)
    cn=cn%26
    if cn==0:cn=26
    return ret+chr(64+cn)

def ColNo2ColRef(cn):
    return colnum2Letter(cn)

import datetime


#helper function: return a set of dates from start_date to end_date to iterate over
def daterange(start_date, end_date):
    for n in range((int ((end_date - start_date).days)) + 1):
        yield start_date + timedelta(n)


def format_number(number, precision=3):
    # build format string
    format_str = '{{:,.{}f}}'.format(precision)

    # make number string
    number_str = format_str.format(number)

    # replace chars
    return number_str.replace(',', 'X').replace('.', ',').replace('X', '.')

def getOtypeByOID(oid):
    rd = genericSelect("select distinct otype from public.excelse_%s where oid='%s'" % (XEDB.DOMAIN, oid))
    if rd:return rd[0].otype.strip()

class CCONT(CONT):
    def __init__(self,cont) -> None:
        self=cont

    def __setattr__(self, name,val):
        iparam=CONT()
        iparam.user="CACHE_ADD"
        setObjectAttribute(iparam,self,name,val,ifdiff=1)
        # Default behaviour
        super(CONT, self).__setattr__(name, val)

def extendAbwOld(r):
    if hasattr(r,"erster_abwesen_heitstag") and hasattr(r,"letzter_abwesen_heitstag"):
        return
    r.erster_abwesen_heitstag=getS(r,"start")
    r.letzter_abwesen_heitstag=getS(r,"ende")

def getAbwArtTag(al,ds):
    ret=[]
    if not al: return ""
    for a in al:
        extendAbwOld(a)
        if not hasattr(a,"erster_abwesen_heitstag"):
            continue
        if ds>=a.erster_abwesen_heitstag and ds<=a.letzter_abwesen_heitstag:
            if a.art!="Elternzeit":
                if "Elternzeit" in ret and "Teilzeit in Elternzeit" in ret:
                    return "Teilzeit in Elternzeit"
                return a.art
            ret.append(a.art)
    if "Praemie Elternzeit" in ret:
        return "Praemie Elternzeit" 
    if "Elternzeit" in ret:
        return "Elternzeit"
    return ""


def namesToVars(names):
    ret = []
    for name in names:
        name = name.lower()
        name = name.strip()
        name = name.replace(" ", "_")
        name = name.replace(",", "_")
        name = name.replace("[", "_")
        name = name.replace("]", "_")
        name = name.replace("(", "_")
        name = name.replace("#", "_")
        name = name.replace("&", "_")
        name = name.replace("%", "pct_")
        name = name.replace(")", "_")
        name = name.replace("\n", "_")
        name = name.replace("\r", "_")
        name = name.replace(":", "")
        name = name.replace(".", "")
        name = name.replace("x000a", "")
        name = name.replace("-", "_")
        name = name.replace("/", "_")
        name = name.replace(u"€", "eur")
        name = name.replace(u"ä", "ae")
        name = name.replace(u"ü", "ue")
        name = name.replace(u"ö", "oe")
        name = name.replace(u"ß", "ss")
        name = name.replace("=", "")
        name = name.replace('"', "")
        name = name.replace("__", "_")
        name = name.replace("__", "_")
        name = name.replace("__", "_")
        while name.endswith("_"):name = name[:-1]
        if name and ord(name[0]) > 65000:name = name[1:]
        ret.append(u"" + name)
    return ret

def getQNRListSorted(htype, qnr):
    gl = getOCACHE()[htype].get(qnr)
    if not gl: return []
    gl2=[]
    gueltig_ab=start=0
    if htype=='abwesenheit_qnr':
        for g in gl:
            if g.start:gl2.append(g)
        gl2.sort(key=lambda x: x.start, reverse=True)
    else:
        for g in gl:
            if g.gueltig_ab:gl2.append(g)
        gl2.sort(key=lambda x: x.gueltig_ab, reverse=True)
    return gl2

def eomday(year, month):
    """returns the number of days in a given month"""
    days_per_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    d = days_per_month[month - 1]
    if month == 2 and (year % 4 == 0 and year % 100 != 0 or year % 400 == 0):
        d = 29
    return d

def daysInYear(year):
    return (datetime.date(year,12,31)-datetime.date(year,1,1)).days+1

class ParseTabText:

    def __init__(self, fname, header=1, seperator="\t", linebreak="",lines=None):
        self.ifile=None
        if lines==None:
            #self.ifile = open(fname)
            self.ifile = open(fname,"r",encoding="utf8")
            for _ in range(1, header):
                _ = self.ifile.readline()
            atts = u"" + self.ifile.readline()
            attl = atts.split(seperator)
            if linebreak != "":
                self.lines = self.ifile.read()
                self.lines = self.lines.split(linebreak)
            else:
                self.lines = None
        else:
            self.lines =lines
            attl=self.lines.pop(0).split(seperator)
        self.seperator = seperator
        self._row = 2
        self.attl = namesToVars(attl)

    def objects(self, abrechnmonat_mm=None):
        while 1:
            if self.lines !=None:
                if self.lines==[]:
                    line=""
                else:
                    line = self.lines.pop(0)
            else:
                line = self.ifile.readline()
                    
            if not line:break
            mdl=matchall('(?<=\")(.*?)(?=\")',line)
            for md in mdl:
                md2=md.replace(self.seperator,"###***###")
                #line=line.replace("\""+md+"\"",md2)
                line=line.replace(md,md2)
            vl = line.split(self.seperator)
            ai = 0
            o = CONT()
            for v in vl:
                if not self.attl[ai]:
                    ai += 1
                    continue
                v = u"" + v.strip()
                if v and ord(v[0]) > 128:v = v[1:]
                v=v.replace("###***###",self.seperator)
                if v.startswith("\"") and v.endswith("\""):
                    v=v[1:-1]
                if "\"" in v:
                    print
                v=v.replace("\"\"","\"")
                setattr(o, self.attl[ai], v)
                ai += 1
            if abrechnmonat_mm and o.abrechnmonat_mm != abrechnmonat_mm:continue
            o._row = self._row
            self._row += 1
            yield o
        if self.ifile:
            self.ifile.close()
        
def getDictFromObject(otype, keya):
    retd = {}
    odict = getOCACHE()[otype]
    for o in odict.values():
        retd[getattr(o, keya)] = o
    return retd

def setNoneAs(inp,toset):
    if not inp:return toset
    return inp

#helper function: return a set of dates from start_date to end_date to iterate over
def daterange(start_date, end_date):
    for n in range((int ((end_date - start_date).days)) + 1):
        yield start_date + datetime.timedelta(n)

def getService(sname, iparam=None, method="GetValues", internal=1):
    omethod = ootype = None
    if iparam:
        ootype = iparam.otype
        omethod = iparam.method
    else:
        iparam = CONT()
        iparam.domain = XEDB.DOMAIN
        iparam.user = "#system"
        iparam.adatumstr = datetime.date.today().strftime("%Y-%m-%d")
        iparam.dtnowstr = datetime.date.today().strftime("%Y-%m-%d")
    iparam.method = method
    if hasattr(iparam, "otypes"):
        ootypes = iparam.otypes
        del iparam.otypes
    else:
        ootypes = None
    iparam.otype = sname
    if internal:
        vals = handle(iparam, 1)
        if hasattr(vals, "retd"):
            vals = list(vals.retd.values())
    else:
        vals = handle(iparam, 0).retd.values()
    if not vals:
        vals = []
    else:
        if sname != "GetFINPersons":
            try:vals = vals[0].odict 
            except:pass
    if ootype:iparam.otype = ootype
    if ootypes:
        iparam.otypes = ootypes
    if omethod:iparam.method = omethod
    return vals

def hasAbwDays(start,end,qnr):
    '''Returns 0 or 1 based on the non-working days defined in ''gehalt'' objects ("salary" objects).
    * 0 means the ``mitarbeiter`` does not have non-working days in the period from "now" until ``end``.
    * 1 means the ``mitarbeiter`` has non-working days in the specified period.

    Inputs are handled as follows:
    * ``qnr`` specifies a ``mitarbeiter`` object ("employee")
    * ``start`` is currently not implemented TODO
    * ``end`` defines the "dateback" in the search loop in all ``gehalt`` objects for the ``mitarbeiter``
    * ``ignore_these`` can be used to define specific entries for "abw_tage" attribute that should be ignored. If nothing is defined, a standard set is used.
    '''    
    gl=getOCACHE()["gehalt_qnr"].get(qnr,[])
    if gl:gl.sort(key=lambda x: setNoneAs(x.gueltig_ab,"0000-00-00"), reverse=False)
    hasabw=0
    for g in gl:
        gab=getS(g,"gueltig_ab")
        abw_tage=getS(g,"abw_tage").lower()
        if gab and gab <=end:
            if abw_tage!="keiner" and abw_tage!="keine" and abw_tage!="": #""can occur if somebody has differing weekly hours but no specific days off -> ignored
                hasabw=1
                break
        if gab and gab>end:
            break
    return hasabw

def extendAbw(r: Any) -> None:
    """
    Extends an 'abwesenheit' record with start and end dates based on specific criteria.

    This function modifies the passed 'abwesenheit' record 'r' directly by setting its
    'erster_abwesen_heitstag' (start of absence) and 'letzter_abwesen_heitstag' (end of absence)
    based on the absence type and related dates (e.g., expected or actual birthdate in case of maternity leave).

    Args:
        r: The 'abwesenheit' record to be modified. It must have 'art', 'start', and 'ende'
           attributes, along with maternity-related dates if applicable.
    
    Modifies:
        r: Directly updates 'erster_abwesen_heitstag' and 'letzter_abwesen_heitstag' if conditions are met.
    """
    art = getS(r, "art")
    r.erster_abwesen_heitstag = getS(r, "start")
    if not r.erster_abwesen_heitstag:
        # If 'start' date is not provided, remove the attribute from 'r'
        del r.erster_abwesen_heitstag
    
    r.letzter_abwesen_heitstag = getS(r, "ende")
    if not r.letzter_abwesen_heitstag:
        # If 'ende' date is not provided, remove the attribute from 'r'
        del r.letzter_abwesen_heitstag
    
    if art == "Mutterschutz":
        # Handling maternity leave specifics
        geb = getD(r, "geplanter_geburtstag_muttersch")
        tgeb = getD(r, "tatsaechlicher_geburtstag_mutt")
        if not tgeb and geb:
            tgeb = geb  # Use planned birthdate if actual birthdate is not provided
        
        if tgeb:
            # Set the 'start' date to 42 days before the birthdate if not already set
            if not hasattr(r, "erster_abwesen_heitstag"):
                e = tgeb - datetime.timedelta(days=42)
                r.erster_abwesen_heitstag = e.strftime(DFMT)
            
            # Set the 'end' date to 56 days after the birthdate if not already set
            if not hasattr(r, "letzter_abwesen_heitstag"):
                e = tgeb + datetime.timedelta(days=56)
                r.letzter_abwesen_heitstag = e.strftime(DFMT)

def templateReplace(mail1,r,eyecatcher="CHR_"):
    if eyecatcher=="CHR_" and XEDB.DOMAIN == "carit":
        eyecatcher="PHR_"
    mailo=mail1
    for a in r.__dict__:
        mailo=mailo.replace(eyecatcher+a,str(getattr(r,a)))
    return mailo

def monthDiff(a,b):
    ym=(b.year-a.year)*12
    mm=b.month-a.month
    return ym+mm

def memAsList(o: Any) -> List[Any]:
    """
    Converts the attributes of an object into a list.

    This function iterates through all attributes of the given object and collects their values into a list.
    It's useful for situations where you need to inspect or serialize the object's current state without needing
    to know the names of the attributes in advance.

    Args:
        o: The object whose attributes are to be converted into a list.

    Returns:
        A list containing the values of all attributes of the object.
    """
    # Access the object's __dict__ attribute to get a dictionary of its attributes and values
    attributes = o.__dict__
    # Initialize an empty list to store the attribute values
    ret = []
    # Iterate through each attribute in the dictionary
    for attribute_name in attributes:
        # Append the value of the attribute to the list
        ret.append(getattr(o, attribute_name))
    return ret

from typing import Any, Dict


def memAsDict(o: Any) -> Dict[str, Any]:
    """
    Converts the attributes of an object into a dictionary.

    This function iterates through all attributes of the given object and maps their names to their values in a dictionary.
    It's useful for serialization, inspection, or when you need to interact with the object's attributes dynamically by name.

    Args:
        o: The object whose attributes are to be converted into a dictionary.

    Returns:
        A dictionary where each key is an attribute name and its corresponding value is the attribute's value.
    """
    # Access the object's __dict__ attribute to get a dictionary of its attributes and values
    attributes = o.__dict__
    # Initialize an empty dictionary to store the attributes as key-value pairs
    ret = {}
    # Iterate through each attribute in the dictionary
    for attribute_name in attributes:
        # Map the attribute name to its value in the return dictionary
        ret[attribute_name] = getattr(o, attribute_name)  # Use 'o' to access the attribute value
    return ret

def handleUpload(req: Any) -> CONT:
    """
    Handles file uploads by processing the uploaded file based on its filename and content.
    
    This function identifies the type of the uploaded file (training data, charging history, payslip)
    and delegates processing to the appropriate handler function. If the file doesn't match any specific
    criteria, it's saved to a determined upload path.
    
    Args:
        req: The request object containing headers, cookies, and files uploaded.
    
    Returns:
        A CONT object with a 'RetCode' indicating the result of the upload operation.
    """
    iparam = CONT()
    addIparamVals(iparam)

    iparam.xsecookie = req.headers.get("phrAccessToken", req.cookies.get("phrAccessToken"))
    
    linf = getLinf(iparam)
    form = req.files
    
    filename = form['file'].filename
    data = form['file'].read()
    
    perf_cards2023 = "perf_cards_2023.zip" == filename
    if perf_cards2023:
        logging.info("Uploading "+filename)
        shared_code.Services.getWorker.readPerfCards2023(iparam, linf, data)
        return


    trainings = "trainings" in filename.lower() and not filename.endswith(".py")
    chargefile = matchall("CarData-ChargingHistory_(.+?)_(.+?).json", filename) or \
                 matchall("BMW-CarData-Ladehistorie_(.+?)_(.+?).json", filename) or \
                 matchall("CarData-Ladehistorie_(.+?)_(.+?).json", filename)
    
    if trainings:
        tmod = reloadService("getTrainings")
        return tmod.handleUpload(iparam, linf, data)
    
    if chargefile:
        return handleChargeFile(iparam, linf, chargefile, data)
    
    psmonth = matchall("0020100.+?([0-9].+).*?.pdf", filename)
    if psmonth:
        tmod = reloadService("splitPayslip")
        return tmod.handleUpload(iparam, linf, data, filename)
    
    uplpath = determine_upload_path(linf, filename)
    logging.info("#########  in directory %s"%os.getcwd())
    with open(uplpath, "wb") as fwrite:
        fwrite.write(data)
    
    ret = CONT()
    ret.RetCode = "ok"
    return ret

def determine_upload_path(linf: dict, filename: str) -> str:
    """
    Determines the upload path for a given file based on the filename and user roles.

    This function checks the user's roles and the filename to determine the appropriate
    upload path. It handles special cases for certain Python files and a specific HTML file,
    placing them in predefined directories. If no special conditions are met, the file is
    directed to a default upload directory.

    Args:
        linf: A dictionary containing information about the user, including their roles.
        filename: The name of the file being uploaded.

    Returns:
        A string representing the determined path where the file should be uploaded.
    """
    uplpath = None

    #linf ist kein dict!  if "hradmin" in linf.get("roles", []):
    if "hradmin" in linf.roles:
        if filename.endswith(".py"):
            if filename in ["WappHandler.py", "sendMail.py"]:
                uplpath = "shared_code/" + filename
            elif "Selbstauskunft" in filename:
                uplpath = "shared_code/Services/" + filename
            else:
                uplpath = "shared_code/Services/" + filename
        if filename in ["BAV_Status.html"]:
            uplpath = "/home/site/wwwroot/dhx/web/web/cf_code_V00048/forms/" + filename

    if not uplpath:
        uplpath = "upload/" + filename

    return uplpath

def getCurrentTimestamp():
    datetime_object = datetime.datetime.now()
    return datetime_object.strftime("%Y-%m-%d %H:%M:%S")

class DELOBJECTS:
    def __init__(self) -> None:
        utcnow=repr(datetime.datetime.utcnow())
        self.tstamp=utcnow[-6:]        
        self.todel=[]
    def add(self, oid):
        self.todel.append(oid)
    def delete(self,dm=None):
        if not dm:dm=self.tstamp
        for oid in self.todel:
            deleteObject(oid,dm)


def unix2TimeString(value,tzone):
    return datetime.datetime.fromtimestamp(value,pytz.timezone(tzone)).strftime('%Y-%m-%d %H:%M:%S')

def setDValue(o,attr,dvalue):
    if not hasattr(o,attr):
        setattr(o,attr,CONT())
    val=getattr(o,attr)
    if not isinstance(val,CONT):#hasattr(val,"__dict__"):
        setattr(o,attr,CONT())
        pval=val
        val=getattr(o,attr)
        val.value=pval
    val.dvalue=dvalue

import inspect


class ADDSUM:
    def __init__(self) -> None:
        self.mdict={}
    def add(self,key,amount):
        self.mdict[key]=self.mdict.get(key,0)+amount
    def get(self,key):
        return self.mdict.get(key,0)

def setAttrCSS(o,attr,css):
    val=getattr(o,attr)
    #if not inspect.isclass(val):
    if not isinstance(val,CONT):#not inspect.isclass(val):
        setattr(o,attr,CONT())
        pval=val
        val=getattr(o,attr)
        val.value=pval
    val.css=css
def addAttrTooltip(o,attr,css):
    val=getattr(o,attr)
    if not isinstance(val,CONT):#not inspect.isclass(val):
        setattr(o,attr,CONT())
        pval=val
        val=getattr(o,attr)
        val.value=pval
    if val.tooltip:
        val.tooltip+="\n"
        val.tooltip+=css
    else:
        val.tooltip=css

def isOverlapping(cejson):
    vals=cejson.values()
    #vals.sort(key=lambda x: x.tstart, reverse=False)
    for (vin,startc),c in cejson.items():
        for cc in vals:
            if cc.tstart==c.tstart: continue
            #neue Regel 02/24: Ladevorgänge mit 0 kWh können ignoriert werden
            if cc.kwh==0 or c.kwh==0:continue
            if cc.tstart<c.tend and c.tstart<cc.tend :
                ret=CONT()
                ret.message="ERROR: overlapping timespans %s until %s and %s until %s"%(c.tstart,c.tend,cc.tstart,cc.tend)  
                return ret


def handleChargeFile(iparam, linf, chargefile, data):
    """
    Handles the processing of a charging file, validating and storing charging events.

    Args:
        iparam: Input parameters for the function.
        linf: User information object.
        chargefile: A list containing information about the charge file.
        data: The raw data of the charging events.

    Returns:
        A CONT object with a RetCode indicating the result of the processing and a message if there were any issues.
    """
    # List of public charging points to be skipped
    PUBLIC_CHARGING_POINTS = ["Schulweg 24, 89081 Ulm"]

    # Fetch skipped chargings for the user
    skippedcd = dictByPNumList("skipped_chargings")

    # Clear skipped chargings cache for the user
    for sc in skippedcd.get(linf.amuser.q_nummer, []):
        del getOCACHE()["skipped_chargings"][sc.oid]
    if linf.amuser.q_nummer in skippedcd:
        del skippedcd[linf.amuser.q_nummer]

    vin = chargefile[0][0]
    ret = CONT()
    ret.RetCode = "ok"

    # Retrieve existing charging events
    chargings1 = getObjects("chargings")
    chargings = {}
    for c in chargings1:
        chargings[(c.vin, c.tstart)] = c

    # Fetch user's location data
    stdo = dictByPNumList("standorthis").get(linf.amuser.q_nummer)
    if not stdo:
        ret.message = "Problem: could not find any standort data for you!"
        return ret

    # Load charging events from data
    cevents = loads(data)

    validcharge = []
    cps2delete = []

    # Validate user location data
    for s in stdo:
        gab = getS(s, "gueltig_ab", "")
        if not s.gueltig_ab:
            continue
        if s.art in ["Heimfahrt"]:
            continue
        if s.gueltig_ab > iparam.dtnowstr:
            continue
        if s.gueltig_bis and s.gueltig_bis < iparam.dtnowstr:
            continue
        validcharge.append("%s, %s" % (s.strasse, s.plz))

    skippeda = {}
    alreadyin = 0
    loadedce = noendtime = skipped_after = 0
    cenr = -1
    cejson = {}

    # Process each charging event
    for ce in cevents:
        cenr += 1
        if ce.publicChargingPoint and ce.publicChargingPoint.__dict__:
            continue
        if ce.chargingLocation and ce.chargingLocation.formattedAddress in PUBLIC_CHARGING_POINTS:
            cpd = CONT()
            cpd.q_nummer = linf.amuser.q_nummer
            cpd.tstart = unix2TimeString(ce.startTime, ce.timeZone)
            cps2delete.append(cpd)
            continue
        if ce.chargingLocation and ce.chargingLocation.municipality == 'Amstetten':
            continue

        # Ensure compatibility with older data formats
        if not ce.calculatedEnergyChargedKwh:
            ce.calculatedEnergyChargedKwh = ce.energyConsumedFromPowerGridKwh
        if not ce.energyChargedKwh:
            ce.energyChargedKwh = ce.energyIncreaseHvbKwh

        ceo = CONT()
        if not ce.chargingLocation:
            ceo.location = "NO_LOCATION"
        else:
            if ce.chargingLocation.formattedAddress:
                ceo.location = ce.chargingLocation.formattedAddress
                ceo.location = ceo.location.replace(", Deutschland", "")
            else:
                ceo.location = ce.chargingLocation.municipality

        # Extract street and postal code from the location
        lsplit = ceo.location.split(",")
        street = lsplit[0].strip()
        hausnr = street.split(" ")
        if len(hausnr) > 1 and hausnr[-1][0].isdigit():
            street = street.replace(hausnr[-1], "").strip()
        if len(lsplit) > 1:
            plzort = lsplit[1].strip()
            plz = plzort.split(" ")[0]
            ceo.location = street + ", " + plz
        else:
            plzort = lsplit[0].strip()
            plz = plzort.split(" ")[0]
            ceo.location = "[NO STREET], " + plz

        # Check if the location is valid for charging
        if ceo.location in validcharge:
            ceo.q_nummer = linf.amuser.q_nummer
            ceo.tstart = unix2TimeString(ce.startTime, ce.timeZone)
            ceo.tend = unix2TimeString(ce.endTime, ce.timeZone) if ce.endTime else ce.endTime
            if not ce.endTime:
                noendtime += 1
                continue
            if ceo.tend[:9] < "2021-07-01":
                continue
            if ceo.tstart > ceo.tend:
                ret.message = "ERROR: invalid timespan included from %s until %s" % (ceo.tstart, ceo.tend)
                return ret
            if (vin, ceo.tstart) in chargings.keys():
                fc = chargings[(vin, ceo.tstart)]
                if fc.q_nummer != ceo.q_nummer:
                    ret.message = "ERROR: the charging event at %s for %s is already registered for another person!" % (ceo.tstart, vin)
                    return ret
                if not fc.calculatedEnergyChargedKwh:
                    setObjectAttribute(iparam, fc, "calculatedEnergyChargedKwh", ce.calculatedEnergyChargedKwh)
                if not fc.energyChargedKwh:
                    setObjectAttribute(iparam, fc, "energyChargedKwh", ce.energyChargedKwh)
                alreadyin += 1
                continue
            ceo.kwh = ce.calculatedEnergyChargedKwh if ce.calculatedEnergyChargedKwh else ce.energyChargedKwh
            ceo.calculatedEnergyChargedKwh = ce.calculatedEnergyChargedKwh
            ceo.energyChargedKwh = ce.energyChargedKwh
            if ceo.kwh == 0:
                continue
            if ceo.tstart[:10] > "2024-02-29" and ceo.tend[:10] < "2024-06-01":
                skipped_after += 1
                continue
            loadedce += 1
            ceo.vin = vin
            cejson[(vin, ceo.tstart)] = ceo
        else:
            if not ce.timeZone:
                ce.timeZone = 'Europe/Berlin'
            getOrAdd(skippeda, ceo.location, []).append(unix2TimeString(ce.startTime, ce.timeZone))

    retover = isOverlapping(cejson)
    if retover:
        return retover

    for ceo in cejson.values():
        setObject(iparam, "chargings", ceo)

    if skippeda:
        ret.message = ("Newly accepted charging events: %ld, events from this file already in: %ld, uncompleted events: %ls\n\n"
                       "Skipped charging locations: %ld\n\n"
                       "Skipped chargings between 2024-03-01 and 2024-05-31: %ld\n\n"
                       "(go to TAB 'SKIPPED CHARGINGS' and click 'reload')" % (loadedce, alreadyin, noendtime, len(skippeda.keys()), skipped_after))
        skippedlocs = skippedcd.get(linf.amuser.q_nummer, [])
        allin = {}
        for s in skippedlocs:
            allin[s.location] = s
        for loc, locl in skippeda.items():
            s = allin.get(loc)
            if s:
                timel = loads(s.timel)
                timec = 0
                for t in locl:
                    if t not in timel:
                        timel.append(t)
                        timec = 1
                if timec:
                    setObjectAttribute(iparam, s, "timel", dumps(timel))
                continue
            ns = CONT()
            ns.q_nummer = linf.amuser.q_nummer
            ns.timel = dumps(locl)
            ns.location = loc
            setObject(iparam, "skipped_chargings", ns, storedb=0)

    logging.info("Chargings to be deleted: \n    " + dumps(cps2delete))
    return ret


def loadPnumDict(otypes):
    for otype in otypes:
        getOCACHE()[otype+"_qnr"]=dictByPNum(otype)

def loadPnumDictList(otypes):
    for otype in otypes:
        getOCACHE()[otype+"_qnr"]=dictByPNumList(otype)

def matchall(regex,s):
    m=re.findall(regex,s,re.U)
    return m

def matchall2(regex,s):
    m=re.findall(regex,s,re.U)
    if m and type(m[0]) is tuple:
        ret=[]
        for me in m[0]:
            ret.append(me)
        return ret
    return m


def getGehaltByDate(qnr, dstring, getfuture=0):
    otypepnr = "gehalt_qnr" if XEDB.DOMAIN == "carit" else "vgehalt_vnr" 
    if otypepnr not in getOCACHE():return None
    if repr(type(dstring)) == "<class 'datetime.date'>":
        dstring = dstring.strftime(DFMT)
    gl0 = getOCACHE()[otypepnr].get(qnr)
    if not gl0:
        return None
    gl=[]
    for g in gl0: #prevents empty gueltig_abs
        if g.gueltig_ab:
            gl.append(g)

    gl.sort(key=lambda x: x.gueltig_ab, reverse=True)
    if getfuture and gl and len(gl) == 1:
        g = gl[0]
        if hasattr(g, "gueltig_ab"):
            if g.gueltig_ab > dstring:
                dstring = "9999-99-99"

    ret = CONT()
    for g in gl:
        if hasattr(g, "gueltig_ab"):
            if dstring >= g.gueltig_ab:
                todel = []
                for f in g.__dict__:
                    try:
                        if f != 'gueltig_ab' and getattr(g, f).replace(' ', '') == '':
                            todel.append(f)
                    except:pass
                for f in todel:
                    delattr(g, f)
                for f in g.__dict__:
                    if not hasattr(ret, f):
                        setattr(ret, f, getattr(g, f))
    return ret

def getActiveLinemanagers(iparam,mad,chkdate):
    ret={}
    for m in mad.values():
        if not maAktiv(None,m,None,chkdate):continue
        lm=getOrgsAsLineManager(m.q_nummer,chkdate)
        if lm:ret[m.q_nummer]=lm
    return ret

def getOrgsAsLineManager(qnum,chkdate):
    """TODO: Add documentation.

    Returns: ret (CONT)
    """
    teams=getOCACHE()["team"]
    abtlgs=getOCACHE()["abteilung"]
    ret=CONT()
    ret.teams=[]
    ret.abtlgs=[]
    for ddict in [teams,abtlgs]:
        for t in ddict.values():
            if t.q_nummer==qnum and chkdate>=getS(t,"gueltig_ab","0000-00-00") and chkdate<=getS(t,"gueltig_bis","9999-99-99"):
                if ddict==teams:
                    ret.teams.append(t.kuerzel)
                if ddict==abtlgs:
                    ret.abtlgs.append(t.kuerzel)
    if not ret.teams and not ret.abtlgs:return None
    return ret

def getMAListByFK(iparam,chkdate):
    mad=dictByPNum("mitarbeiter")
    lms=getActiveLinemanagers(iparam,mad,chkdate)
    ret={}
    for m in mad.values():
        if not maAktiv(None,m,None,chkdate):
            continue
        v=getVertragByDate(m.q_nummer,chkdate)
        if not v:
            continue
        if m.q_nummer in lms.keys() and lms[m.q_nummer].abtlgs:
            getOrAdd(ret,"q131901",[]).append(m.q_nummer)
            continue
        if m.q_nummer in lms.keys():
            for qnr, lm in lms.items():
                if v.abteilung in lm.abtlgs:
                    getOrAdd(ret,qnr,[]).append(m.q_nummer)
                    continue
            continue
        for qnr, lm in lms.items():
            if v.team in lm.teams:
                getOrAdd(ret,qnr,[]).append(m.q_nummer)
                continue
    return ret

def JForm2ToObj(fbl):
    ret=CONT()
    for f in fbl:
        setattr(ret,f.name,f.value)
    return ret

def getHistoryByDate(htype,qnr,dstring,cnd_field=None,cnd_value=None,getfuture=0):
    if htype not in getOCACHE():return None
    dtype=repr(type(dstring))
    if dtype in ["<class 'datetime.date'>","<class 'datetime.datetime'>"]:dstring=dstring.strftime(DFMT)
    gl=getOCACHE()[htype].get(qnr)
    if not gl:
        return None
    if 0:
        for g in gl:
            if not g.gueltig_ab:
                g.gueltig_ab="0000-00-00"
        gl.sort(key=lambda x: x.gueltig_ab, reverse=True)
    gl.sort(key=lambda x: setNoneAs(x.gueltig_ab,"0000-00-00"), reverse=True)

    if getfuture and gl and len(gl) == 1:
        g = gl[0]
        if g.gueltig_ab and g.gueltig_ab > dstring:
            dstring = "9999-99-99"

    ret=CONT()
    blockl=[]
    for g in gl:
        if cnd_field:
            if cnd_field=="art" and htype=="standort_qnr" and not getS(g,cnd_field):
                if "Erstw." != cnd_value:
                    continue
            elif getS(g,cnd_field) != cnd_value:
                continue
        if hasattr(g, "gueltig_ab"):
            if dstring>=g.gueltig_ab:
                for f in g.__dict__:
                    if f in blockl:
                        continue
                    if not hasattr(ret, f):
                        val=getattr(g, f)
                        if val:
                            setattr(ret, f, val)
                if "strasse" in g.__dict__: blockl.append("adresszusatz")

    return ret

def safe_delattr(obj, attrname):
    if hasattr(obj, attrname):
        delattr(obj, attrname)

def makeXLSDate(value):
    ret = CONT()
    cell = CONT()
    cell.value = value
    cell.field = "datum"
    validateValue(None, cell, ret)
    return cell.value


def getStandortByDate(qnr,dstring,cnd_field=None,cnd_value=None):
    """TODO: Add documentation.

    Returns: getHistoryByDate("standort_qnr",qnr,dstring,cnd_field,cnd_value)
    """
    return getHistoryByDate("standort_qnr",qnr,dstring,cnd_field,cnd_value)

def reloadHistory(otypes):
    dummy=getOCACHE()
    for ot in otypes:
        if ot in ["vertrag","standort"]:
            othis="vertragshis" if ot=="vertrag" else ot+"his"
            getOCACHE()[ot+"_qnr"]=dictByPNumList(othis)
        else:
            getOCACHE()[ot+"_qnr"]=dictByPNumList(ot)

def getVertragByDate(qnr,dstring,cnd_field=None,cnd_value=None,getfuture=0):
    """TODO: Add documentation.

    Returns: getHistoryByDate("vertrag_qnr",qnr,dstring,cnd_field,cnd_value)
    """
    return getHistoryByDate("vertrag_qnr",qnr,dstring,cnd_field,cnd_value,getfuture=getfuture)

def getWstMax(vart):
    if vart in ["Doktorand", "Studienabschlussarbeit", "Praktikant", "Werkstudent"]: wstmax = 35
    else: wstmax = 40
    return wstmax

def getOrgInfo(qnum,CHKDATE,r=None):
    v=getVertragByDate(qnum,CHKDATE)
    if r:ret=r
    else:
        ret=CONT()
    ret.team=None if not v else v.team
    ret.abteilung=None if not v else v.abteilung
    return ret

def getTeam(qnum,CHKDATE):
    v=getVertragByDate(qnum,CHKDATE)
    if not v:
        v=getOCACHE()["planzak_qnr"].get(qnum)
    if not v:
        v=getOCACHE()["planzak_qxnr"].get(qnum)
    if not v:
        return None
    return v.team

def getAbteilung(qnum,CHKDATE):
    v=getVertragByDate(qnum,CHKDATE,getfuture=1)
    if not v:
        v=getOCACHE()["planzak_qnr"].get(qnum)
    if not v:
        v=getOCACHE()["planzak_qxnr"].get(qnum)
    if not v:return None
    return v.abteilung

def getOrgByDate(ma,orgtyp,kuerzel,dstring1):
    """TODO: Add documentation.

    Returns: g (any)
    """
    if ma:
        austritt=getS(ma,"austrittsdatum","9999-99-99",True) #True means that empty string has to be replaced as if not existent
        dstring=min(dstring1,austritt)
    else:
        dstring=dstring1
    #dummy=getOCACHE()["%s_kuerzel"%orgtyp]
    if not kuerzel:
        return
    if 0 and "-" in kuerzel and kuerzel.endswith("DR"):
        kuerzel=kuerzel[:-2]
        orgtyp="abtlg"
    orgl=getOCACHE()["%s_kuerzel"%orgtyp].get(kuerzel)
    if not orgl: return None
    try:
        orgl.sort(key=lambda x: x.gueltig_ab, reverse=True)
    except:
        for g in orgl:
            if getS(g, "gueltig_ab","")=="":
                g.gueltig_ab="1970-01-01"
        orgl.sort(key=lambda x: x.gueltig_ab, reverse=True)
        
    for g in orgl:
        if hasattr(g, "gueltig_ab"):
            if dstring>=g.gueltig_ab:
                break
    if g.kuerzel in ["NoTeam","#NV","NoAbtlg"]:
        return None
    return g


def dictByPNum(otype):
    ret={}
    if otype in getOCACHE():
        for o in getOCACHE()[otype].values():
            if not o.q_nummer:
                continue
            ret[o.q_nummer]=o
    return ret

def dictByPNumList(otype):
    ret={}
    if otype in getOCACHE():
        for o in getOCACHE()[otype].values():
            if not o.q_nummer:
                continue
            getOrAdd(ret,o.q_nummer,[]).append(o)
    return ret

def dictByAttrList(otype,attr):
    ret={}
    if otype in getOCACHE():
        for o in getOCACHE()[otype].values():
            val=getattr(o,attr)
            if not val:continue
            getOrAdd(ret,val,[]).append(o)
    return ret

def dictByAttr(otype,attr):
    """TODO: Add documentation.

    Returns: ret (dict)
    """
    ret={}
    if otype in getOCACHE():
        for o in getOCACHE()[otype].values():
            if hasattr(o,attr):
                val=getattr(o,attr)
                if not val:continue
                ret[val]=o
    return ret

def copyAttribsFromUIGRID(m,g,uigrid):
    if not g:
        return
    for f in uigrid["fields"]:
        fname=f["name"]
        if fname in g.__dict__:
            val=getattr(g, fname)
            setattr(m, fname, val)

def copyAttribs(m,g,repl_nl=0):
    if not g:
        return
    for a in g.__dict__:
        val=getattr(g, a)
        if repl_nl:
            val=val.replace("\n","<br>")
        setattr(m, a, val)

def setObject(iparamcall,otype,o,attrib=None,aval=None,storedb=1,nohistory=0):
    '''Sets a new object or updates an existing object.'''
    if not storedb:
        if not getOCACHE().get(otype):getOCACHE()[otype]={}
        thisocache=getOCACHE().get(otype)
        co=None
        if o.oid:
            co=thisocache.get(o.oid)
        else:
            o.oid="C%06ld"%(len(thisocache)+1)
        if not co:co=CONT()
        copyAttribs(co,o)
        thisocache[co.oid]=co
        return o
    iparam=CONT()
    copyAttribs(iparam, iparamcall)
    iparam.method="SetValues"
    iparam.otype=otype
    iparam.noextend=1
    iparam.cells=[]
    if 0 and hasattr(o, "oid"):
        oc=getObject(otype, o.oid)
    else:
        oc=None
    #for o in ol:
    if attrib:
        setattr(o,attrib,aval)
        al=[attrib] 
    else:
         al=o.__dict__
    for a in al:
        if a!="oid":
            c=CONT()
            c.field=a
            c.value=getattr(o, a)
            dd=repr(type(c.value))
            if dd=="<type 'time'>":
                val=repr(c.value)
                ll=val.split(".")
                c.value=ll[2][0:4]+"-"+ll[1]+"-"+ll[0][8:10]
            if dd=="<class 'decimal.Decimal'>":
                c.value=float(c.value)
            if dd in ["<type 'instance'>","<type 'list'>","<type 'dict'>"]:
                c.field+=":JSON"
                c.value=base64.b64encode(dumps(c.value))
            if hasattr(o, "oid"):
                c.oid=o.oid
            else: c.oid="R1"
            if oc and hasattr(oc,a):
                if getattr(oc,a)==c.value:continue
            iparam.cells.append(c)
    olist,ret=handleSetValues(iparam,internal=1,nohistory=nohistory)
    if olist:
        return olist[0]
    else:
        return None

def intersect(list1,list2):
    return [x for x in list1 if x in list2]    


def genericSelect(sql: str, verbose: int = 1, typeconv: int = 1) -> list:
    """
    Executes a SQL query and returns the results as a list of CONT objects.

    This function connects to a database using a predefined connection (XEDB.DBConn),
    executes a given SQL query, and processes the results. Each row from the query result
    is converted into a CONT object with attributes corresponding to the column names. 
    The function can optionally perform type conversion based on the column's type code,
    and it supports verbose output of the SQL query.

    Parameters:
    - sql (str): The SQL query to be executed.
    - verbose (int, optional): If set to 1 (default), the SQL query is printed before execution.
    - typeconv (int, optional): If set to 1 (default), performs type-specific conversions 
                                for known type codes (e.g., formatting datetime objects or 
                                stripping strings). If set to 0, no type conversion is performed.

    Returns:
    - list: A list of CONT objects, each representing a row from the query result. Attributes
            of CONT objects correspond to the SQL query's column names.

    Note:
    - The function assumes the existence of a global database connection (XEDB.DBConn).
    - CONT is a custom class used to dynamically create objects based on query results.
    - Type conversion is limited to specific type codes and might need adjustment based on
      the database schema or requirements.
    """
    if verbose:
        print(sql)
    ret = []
    cursor = XEDB.DBConn.cursor()
    cursor.execute(sql)
    
    records = cursor.fetchall()
    for rs in records:
        r = CONT()
        cn = -1
        for col in cursor.description:
            cn += 1
            if rs[cn] is None:
                continue
            if typeconv:
                if col.type_code in [23]:  # Assuming 23 is an integer type
                    setattr(r, col.name, rs[cn])
                elif col.type_code in [1114]:  # Assuming 1114 is a datetime type
                    setattr(r, col.name, rs[cn].strftime("%Y-%m-%d %H:%M:%S"))
                else:
                    # For unrecognized types, attempt to strip the string if conversion is enabled
                    setattr(r, col.name, rs[cn].strip())
            else:
                # If type conversion is not desired, set attribute without conversion
                setattr(r, col.name, rs[cn])
        ret.append(r)
    
    return ret


def setForeignValue(r,o,loc_field,for_type,for_oid,for_field=None):
    if for_field==None: for_field=loc_field
    dummy=getOCACHE()[for_type]
    if 1 or hasattr(r, "qnr"):
        try:
            setattr(r,loc_field,getattr(getOCACHE()[for_type][getattr(o,for_oid)], for_field))
            return 1
        except:
            setattr(r,loc_field,"#NV")
            return 0

class TStamp:
    def __init__(self,sstr=None):
        if sstr: print(sstr)
        self.start = datetime.datetime.now()
    def gap(self):
        end = datetime.datetime.now()
        gap=end-self.start
        self.start=end
        return gap.seconds*1000+gap.microseconds/1000

def copy_attributes(target_obj, source_obj, replace_newlines=0):
    """
    Copy attributes from the source object to the target object.
    Improved version of copyAttribs (more readable, more comments)

    :param target_obj: The object to which attributes will be copied.
    :param source_obj: The object from which attributes will be copied.
    :param replace_newlines: If set to 1, replace newline characters with '<br>' in the values.
    """
    
    # If the source object is None or doesn't have attributes, return without doing anything.
    if not source_obj:
        return
    
    # Iterate over attributes of the source object
    for attr_name in source_obj.__dict__:
        attr_value = getattr(source_obj, attr_name)
        
        # If replace_newlines is set, replace newline characters in attribute value
        if replace_newlines:
            attr_value = attr_value.replace("\n", "<br>")
        
        # Set the attribute on the target object
        setattr(target_obj, attr_name, attr_value)


def runDBStatement(sql: str, verbose: int = 0) -> Optional[Any]:
    """
    Executes a given SQL statement on the database and optionally prints the SQL statement.

    Parameters:
    - sql (str): The SQL statement to be executed.
    - verbose (int): If set to a non-zero value, the SQL statement will be printed to stdout. Defaults to 0.
    
    Returns:
    - Optional[Any]: The first column of the first row of the result set, if any results are returned. Otherwise, None.
    """
    
    # Print the SQL statement if verbose mode is enabled.
    if verbose:
        print(sql)
    
    # Establish a database cursor from the connection.
    cursor = XEDB.DBConn.cursor()
    
    # Execute the SQL statement.
    cursor.execute(sql)
    
    # Fetch the first row of the results if there are any, else return the status message instead.
    result = None
    if cursor.description:
        result = cursor.fetchone()[0]
    else:
        result=cursor.statusmessage
    
    # Commit the transaction to the database to ensure changes are saved.
    XEDB.DBConn.commit()
    
    # Return the query result (if any) or None.
    return result


def getObjFromDB(sql,lasttidinp=0,nocache=0):
    retd={}
    lasttid=0
    otypes=[]
    cursor = XEDB.DBConn.cursor()
    cursor.execute("select "+sql)
    records = cursor.fetchall()
    if nocache:
        oc=CONT()
    if records:
        colpos={}
        cn=-1
        for col in cursor.description:
            cn+=1
            colpos[col.name]=cn
        for rs in records:
            lid= rs[colpos["tid"]]
            if lid>lasttid:lasttid=lid
            field= rs[colpos["field"]].strip()
            oid= rs[colpos["oid"]].strip()
            otype= rs[colpos["otype"]].strip()
            if otype not in otypes:
                otypes.append(otype)
            val_string= rs[colpos["val_string"]].strip()
            if field.endswith(":JSON"):
                field=field[:-5]
                val_string=loads(base64.b64decode(val_string))
            ret=getOrAdd(retd,otype,CONT())
            if not ret.odict:
                ret.odict={}
            if not nocache:
                oc=objFromCache(otype,oid,1)
            setattr(oc,field,val_string)
            ret=getOrAdd(retd,otype,CONT())
            if not hasattr(ret,"odict"):
                ret.odict={}
            ret.odict[oid]=oc

    if lasttid==0: lasttid=lasttidinp

    return (retd,lasttid,otypes)

def loadCache():
    return getObjFromDB("distinct on (otype, oid, field) otype, oid, field, val_string, tid FROM excelse_%s where tid > %ld and status is null %s order by otype, oid desc, field, tid desc"%(XEDB.DOMAIN,0,""),0)

def refreshCache(iparam):
    (retd,XEDB.lasttid)=getObjFromDB("distinct on (otype, oid, field) otype, oid, field, val_string, tid FROM excelse_%s where otype='%s' and tid > %ld and status is null %s order by otype, oid desc, field, tid desc"%
                                               (XEDB.DOMAIN,iparam.otype,XEDB.lasttid,iparam.extracond),XEDB.lasttid)

def cleanLogins():#alte Logins eines Users rauswerfen
    delo=DELOBJECTS()
    logins={}
    ret=[]
    for li in getObjects("sys_login"):
        pid=li.amuser.real_v_num if li.amuser.real_v_num else li.amuser.v_num
        oldli=logins.get(pid)
        if oldli:
            if oldli.tstamp<li.tstamp:
                delo.add(oldli.oid)
                logins[pid]=li
            else:
                delo.add(li.oid)
        else:
            logins[pid]=li
    ret=delo.todel
    delo.delete("OLDLIS")
    return ret

def refreshCacheAll(iparam):
    lbefore=XEDB.lasttid
    try:
        (retd,XEDB.lasttid,otypes)=getObjFromDB("distinct on (otype, oid, field) otype, oid, field, val_string, tid FROM excelse_%s where tid > %ld and status is null and not otype in ('#sysexception') order by otype, oid desc, field, tid desc"%
                                                (XEDB.DOMAIN,XEDB.lasttid),XEDB.lasttid)
    except:
        try:
            XEDB.DBConn.commit()
        except:
            XEDB.DBConn = psycopg2.connect(XEDB.CONSTRING%XEDB.DBPASSWORD)
            XEDB.DBConn.commit()
        (retd,XEDB.lasttid,otypes)=getObjFromDB("distinct on (otype, oid, field) otype, oid, field, val_string, tid FROM excelse_%s where tid > %ld and status is null and not otype in ('#sysexception') order by otype, oid desc, field, tid desc"%
                                                (XEDB.DOMAIN,XEDB.lasttid),XEDB.lasttid)

    if lbefore!=XEDB.lasttid:
        print("### Transactions loaded into cache: %ld - object types affected:%s"%(XEDB.lasttid-lbefore,repr(sorted(otypes)))                                         )
        if not XEDB.AZURE and "sys_login" in retd:
            cleaned=cleanLogins()
            for li in retd["sys_login"].odict.values():
                if li.oid in cleaned:continue
                XEDB.COOKIES[li.xsecookie]=li
    if not XEDB.CACHE_LOADED:
        XEDB.CACHE_LOADED=1


def objFromCache(otype,oid,gen=0):
    odict=getOrAdd(getOCACHE(),otype,{})
    r=odict.get(oid)
    if gen and not r:
        r=CONT()
        r.oid=oid
        odict[oid]=r
    return r

def reloadService(sname):
    modname='shared_code.Services.%s'%(sname)
    
    try:
        smodule=eval(modname)
        if 1:# and XEDB.ENVIRONMENT in ["TEST"]:
            logging.info(f"Reloading module {modname}.")
            importlib.reload(smodule)
    except:
        etext=traceback.format_exc()
        if "module 'shared_code.Services' has no attribute '%s'"%sname not in etext:
            logging.warning(etext)
        else:
            logging.info(f"Loading {modname}.")
        smodule=importlib.import_module(modname)
    return smodule

def isTeamlead(mitarbeiter: object, date):
    """Checks if a person is a teamlead.

    Returns: state (0 or 1)
    """
    state=0
    relative_team=mitarbeiter.team
    relative_teamlead=getOrgByDate(None,"team",relative_team,date).q_nummer
    if relative_teamlead==mitarbeiter.q_nummer:
        state=1
    return state

def isDirectReport(my_abteilung: str, mitarbeiter: object):
    """Checks if a person is a direct report.

    Returns: state (0 or 1)
    """
    state=0
    relative_team=mitarbeiter.team
    relative_abteilung=mitarbeiter.abteilung[0:4]
    direct_report_short=relative_team[4:6]
    if relative_abteilung==my_abteilung and direct_report_short=="DR":
        state=1
    
    return state

def toCamelCase(ss):
    sl=ss.split("_")
    ret=""
    for s in sl:
        ret+=s[0].upper()+s[1:].lower()
    return ret

def saveDBFile(iparam,fc,filename):
    import traceback
    try:
        fc64=base64.b64encode(fc).decode()
        sql="INSERT INTO excelse_%s_files (xseuser,otype,val_string,oid,field) VALUES('%s','%s','%s','%s','%s') RETURNING tid;"%(iparam.domain,iparam.user,"file",fc64,"R0","content",)
        tid=runDBStatement(sql,verbose=0)
        noid="I%09ld"%tid
        sql="UPDATE excelse_%s_files set oid = '%s' where tid=%ld;"%(iparam.domain,noid,tid)
        runDBStatement(sql)
        sql="INSERT INTO excelse_%s_files (xseuser,otype,val_string,oid,field) VALUES('%s','%s','%s','%s','%s') RETURNING tid;"%(iparam.domain,iparam.user,"file",filename,noid,"filename",)
        runDBStatement(sql,1)
        return noid
    except:
        print(traceback.format_exc())
        return None

class DBFile:
    def __init__(self,fname, oid) -> None:
        self.fname=fname
        self.oid=oid

def getDBFile(iparam,fname):
    #rs=genericSelect("select val_string,oid from excelse_carit_files where field='filename' and val_string='%s'"%fname)
    rs=genericSelect("select distinct on (otype, oid, field) otype, oid, field, val_string, tid FROM excelse_%s_files where status is null and field='filename' and val_string='%s' order by otype, oid desc, field, tid desc"%(XEDB.DOMAIN,fname))
    if not rs:
        return None
    #fo=getObjFromDB("distinct on (otype, oid, field) otype, oid, field, val_string, tid FROM excelse_%s_files where tid > %ld and status is null %s and oid='%s' order by otype, oid desc, field, tid desc"%(XEDB.DOMAIN,0,"",rs[0].oid),0,nocache=1)
    fo=DBFile(fname, rs[0].oid)
    return fo

def downloadDBFile(iparam,fname):
    fo=getDBFile(iparam,fname)
    ret=CONT()
    if not fo:
        ret.RetCode="File not found!"
        return ret

    rs=genericSelect("select val_string from excelse_carit_files where field='content' and oid='%s'"%fo.oid)
    ret.gridinfos=[]
    ret.filename=fname
    ret.filelength=len(rs[0].val_string)
    ret.stream=rs[0].val_string
    iparam.download=1
    return ret
    

def handleDownloadGen(iparam,olist):
    linf=getLinf(iparam)
    if iparam.otype=='pofomgmnt' and iparam.selected:
        hrlead="hrlead" in linf.roles
        ret=CONT()
        #ret.retCode="Not yet activated"
        #return ret
        po=getObject(iparam.otype+iparam.pofoyear[2:],iparam.selected[0])
        if po.reviewer!=linf.user and not hrlead:
            ret.RetCode="Not Authorized"
            return ret
        fname="POFO2023DL/%s_POFO2023.zip"%po.q_nummer
        return downloadDBFile(iparam,fname)
        


    ret=CONT()
    safe_delattr(iparam,"download")
    #del iparam.download
    if not olist:
        sdict=getService(iparam.otype,iparam)
    #xls=MEXCEL("")
    xls=MEXCEL("Templates/gen_template.xlsx")
    #t=xls.getTable()


    #ws=xls.wb.worksheets["raw_data"]

    for sh in xls.wb:
        for pivot in sh._pivots:
            pivot.cache.refreshOnLoad = True

    ws=xls.wb["raw_data"]

    # add column headings. NB. these must be strings

    headers=[]
    attributes=[]
    col=0

    colbreak=0

    if not olist:
        smodule=XEDB.DYNMODS[iparam.otype]

        if hasattr(smodule,"getUIGRID"):
            reload(smodule)
            gridinfo=smodule.getUIGRID(iparam)
        else:
            if hasattr(smodule,"UIGRID"):
                gridinfo=smodule.UIGRID

        for uif in gridinfo["fields"]:
            if not uif["text"]:
                continue
            col+=1
            if colbreak and col > colbreak:
                break

            headers.append(uif["text"])
            attributes.append(uif["name"])
            ws.column_dimensions[colnum2Letter(col)].width = uif["width"]+0.01
    else:
        attribs=[]
        for o in olist:
            for a in o.__dict__:
                appendUnique(attribs,a)
        for a in attribs:
            col+=1
            if colbreak and col > colbreak:
                break
            headers.append(a)
            attributes.append(a)
            ws.column_dimensions[colnum2Letter(col)].width = 15
        attributes.sort()
        headers.sort()
    print(headers)
    ws.append(headers)
    row=1
    if not olist:olist=sdict.values()
    for o in olist:
        row+=1
        col=0
        for attr in attributes:
            col+=1
            if colbreak and col > colbreak:
                break
            if hasattr(o,attr) and getattr(o,attr)!=None:
                val=getattr(o,attr)
                if isCont(val):
                    val=val.value
                if  isinstance(val, str) and val.startswith("<button>"):continue
                ws.cell(row,col).value=val

    tab = Table(displayName="raw_data", ref="A1:%s%ld"%(colnum2Letter(len(headers)),row))

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    '''
    Table must be added using ws.add_table() method to avoid duplicate names.
    Using this method ensures table name is unque through out defined names and all other table name. 
    '''
    ws.add_table(tab)


    ret.gridinfos=[]

    ret.filename="payarahr_%s.xlsx"%iparam.otype
    xls.saveAsStream(ret)
    iparam.download=1
    return ret



def getOldestObjectTStamp(oid):
    """Because the tid is used to generate the OID in turn"""
    rd = genericSelect("select MIN(tstamp) from public.excelse_%s where oid='%s'" % (XEDB.DOMAIN, oid),verbose=0)
    if rd:return rd[0].min



def getFunc(iparam, gset="get"):
    """
    Dynamically selects and returns a service function based on the input parameters.
    Args:
        iparam: An object that contains the method to be executed and possibly an 'otype' attribute,
                which specifies the type of object the method operates on. This object may also
                require modification if its 'otype' attribute ends with a digit followed by a
                non-digit character.
        gset: A string indicating the operation mode, defaults to "get". It is used to construct
              the service method name dynamically, particularly when handling 'GetValues' or
                'SetValues' methods.
    Returns:
        A function object from the dynamically loaded module that corresponds to the requested
        service method.
    This function dynamically determines the service function to be executed based on the 'method'
    attribute of the 'iparam' object. It modifies 'iparam.otype' under specific conditions,
    constructs the function name to be called, and dynamically imports the required module
    before returning the relevant function.
    """

    # Check if 'iparam' has 'otype' attribute ending with a digit followed by a non-digit, modify if true
    if hasattr(iparam, "otype") and iparam.otype[-1].isdigit() and not iparam.otype[-2].isdigit():
        iparam.org_otype = iparam.otype  # Backup original 'otype'
        iparam.otype = iparam.otype[:-1]  # Remove last character from 'otype'

    # Determine the service name and function name based on the method
    if iparam.method in ["GetValues", "SetValues"]:
        otype2 = iparam.otype
        sname = gset + toCamelCase(otype2)  # Construct service name dynamically
    else:
        sname = otype2 = iparam.method

    fname = 'shared_code.Services.%s.%s' % (sname, sname)  # Construct function name

    # Special handling for XLSDownload method
    if iparam.method == "XLSDownload":
        if service_has_individual_download_function(sname):
            sname = gset + toCamelCase(iparam.otype)
            smodule = reloadService(sname)
            fname = "handleDownload"
            fname = 'shared_code.Services.%s.%s' % (sname, fname)
        else:
            fname = "handleDownloadGen"
            otype2 = iparam.otype
            sname = gset + toCamelCase(otype2)

    # Attempt to reload service module, fallback to generic service if failed
    try:
        smodule = reloadService(sname)
    except:
        smodule = reloadService("getGenData")  # Fallback service
        if otype2 not in smodule.UIGRIDS:
            reloadService(sname)  # Attempt reload original service if not in UIGRIDS
        fname = 'shared_code.Services.getGenData.getGenData'

    func = eval(fname)  # Dynamically evaluate and get function object

    XEDB.DYNMODS[otype2] = smodule  # Update dynamic modules dictionary

    return func  # Return the dynamically determined function object

def service_has_individual_download_function(service_name: str):
    path = 'shared_code.Services'
    if service_name in ["trainings","forecast","feedback_consol"]:
        return True
    else:
        return False

def email2name(em):
    em=em.strip().lower()
    em=em.split("@")[0]
    em=em.split(".")
    ret=CONT()
    if len(em)>1:
        ret.vorname=em[0][0].upper()+em[0][1:]
        ret.nachname=em[-1][0].upper()+em[-1][1:]
        ret.name="%s, %s"%(ret.nachname,ret.vorname)
    else:
        ret.vorname=ret.nachname=ret.name=em[0][0].upper()+em[0][1:]
    return ret

def getLineManagers(cdate):
    ret={}
    for ot in ["team","abteilung","gmbh"]:
        #print
        for o in getObjects(ot):
            if not o.q_nummer:continue
            gab=getS(o,"gueltig_ab")
            if not gab:continue
            gbis=getS(o,"gueltig_bis","9999-99-99")
            if cdate<=gbis and cdate >=gab:
                no=CONT()
                no.q_nummer
                no.type=ot
                no.kuerzel=o.kuerzel
                getOrAdd(ret,o.q_nummer,[]).append(no)
    return ret


def setFValue(r,o,loc_field,for_field=None,valnf="#NV"):
    if for_field==None: for_field=loc_field
    try:
        setattr(r,loc_field,getattr(o, for_field))
        return 1
    except:
        if valnf==None:return 0
        setattr(r,loc_field,valnf)
        return 0


def getObject(otype,oid):
    """Returns an object of the type ``otype`` with the specific Object ID ``oid`` from the OCACHE.
    """
    return getOCACHE().get(otype,{}).get(oid)

def getObjects(otype,olist=None):
    '''* Returns a ``list`` of objects of the type ``otype`` from the ``OCACHE``.
    * If you set an ``olist`` it only returns one object: The one specified by the
    [0] index of the ``olist``.
    * TODO: Clarify why this only returns one object when the param is called "olist"
    '''
    ocache=getOCACHE()
    odict=ocache.get(otype)

    if not odict:
        return []
    if olist:
        return [odict[olist[0].oid]]
    
    return odict.values()

def getActiveMitarbeiterDictByQnr(iparam):
    '''* Returns a ``dict`` of active ``mitarbeiter`` ("employee") objects` from the ``OCACHE``.
    '''
    ret={}
    ocache=getOCACHE()
    odict=ocache.get("mitarbeiter")

    if not odict:
        return ret
    else:
        for employee in odict.values():

            #Check for q_nummer (personal ID for employee)
            if not employee.q_nummer:
                continue

            #Check if the employee is active
            ma_aktiv=maAktiv(iparam,employee)
            if not ma_aktiv:
                continue

            #If the employee has a q_nummer and is active we can write them to the dict
            #Usually only an employee with a q_nummer can be active - but I have seen things...
            ret[employee.q_nummer]=employee

    
    return ret

def getObjectDict(otype,olist=None):
    '''* Returns a ``dict`` of objects of the type ``otype`` from the ``OCACHE``.
    * If you set an ``olist`` it only returns one object: The one specified by the
    [0] index of the ``olist``.
    * TODO: Clarify why this only returns one object when the param is called "olist"
    '''
    ocache=getOCACHE()
    odict=ocache.get(otype)

    if not odict:
        return {}
    if olist:
        return {olist[0].oid:odict[olist[0].oid]}
    
    return odict

def getEnterTabs(s):
    ret=[]
    for l in s.split("\n"):
        ret.append(l.split("\t"))
    return ret

def buildFakedUser(iparam):
    linf=CONT()
    linf.amuser=CONT()
    linf.roles=["changeuser","hreditor","hradmin"]#,"hradmin"]#["changeuser","hradmin"]
    linf.amuser.sub="qx60276"
    linf.amuser.q_nummer=linf.amuser.sub
    linf.user=linf.amuser.sub
    XEDB.COOKIES[iparam.xsecookie]=linf
    return linf


def getLinf(iparam):
    if not hasattr(iparam,"xsecookie"):
        return None
    linf=XEDB.COOKIES.get(iparam.xsecookie)
    if 0 and not linf and XEDB.ENVIRONMENT in ["TEST"]:
        iparam.xsecookie="FAKEDLOCAL"
        return buildFakedUser(iparam)
    return linf

DFMT=u"%Y-%m-%d"

def setSuggestedValue(iparam, r, otype, oid, field, val, noval=None):  # noval auf "" zu setzen setzt den Value, wenn "" drin steht
    oc = objFromCache(otype, oid)
    if hasattr(oc, field): 
        if noval == None:return
        if getattr(oc, field) != noval:return
    setattr(oc, field, val)
    setattr(r, field, val)
    sql = "INSERT INTO excelse_%s (xseuser,otype,val_string,oid,field) VALUES('%s','%s','%s','%s','%s');" % (iparam.domain, "#syscomp", otype, val, oid, field,)
    print(sql)
    runDBStatement(sql)

BAVMAP=None

class MAPBAV(CONT):
    def __init__(self,chkdate=None) -> None:
        super().__init__()
        self.bmap=dictByAttrList("bavmap","stufe")
        self.mapit={}
        if not chkdate:
            dtnow=datetime.date.today()
            self.chkdate=dtnow.strftime(DFMT)
        else:
            self.chkdate=chkdate
        for bl in self.bmap.values():
            bl.sort(key=lambda x: x.gueltig_ab, reverse=True)
            for bm in bl:
                if self.chkdate >=bm.gueltig_ab:
                    self.mapit[bm.stufe]=bm
                    break
    def map(self,stufe):
        return self.mapit.get(stufe)

def mapBAV(o):
    global BAVMAP
    if not BAVMAP:
        dtnow=datetime.date.today()
        dtnowstr=dtnow.strftime(DFMT)
        BAVMAP={}
        for b in getObjects("bavmap"):
            if dtnowstr<b.gueltig_ab:continue
            bm=BAVMAP.get(b.stufe)
            if bm and bm.gueltig_ab>b.gueltig_ab:continue
            BAVMAP[b.stufe]=b

    sbav=getN(o,"bav")
    if sbav:return sbav
    plevel=getS(o,"prof_level") #if not kstufe else ""
    kstufe=getS(o,"aktuelle_einstufung") if not plevel else ""
    fnd=BAVMAP.get(plevel)
    if fnd:return int(fnd.betrag)
    fnd=BAVMAP.get(kstufe)
    if fnd:return int(fnd.betrag)
    if "SL" in plevel:
        if "3" in kstufe or "4" in plevel: return 105 
        if "2" in kstufe or "3" in plevel: return 80 
        if "1" in kstufe or "2" in plevel: return 55
        return 0
    if "7" in plevel: return 170 
    if "6" in plevel: return 170 
    if "5" in kstufe or "5" in plevel: return 170 
    if "4" in kstufe or "4" in plevel: return 135 
    if "3" in kstufe or "3" in plevel: return 105 
    if "keine" in kstufe: return 0 
    if "2" in kstufe or "2" in plevel: return 80 
    if "1" in kstufe or "1" in plevel: return 55
    if "ohne" in kstufe: return 80 
    if "ohneOM" in kstufe: return 55
    return 0


def getFKHistory(orgtyp,org):
    kuerzel=getS(org,"kuerzel")
    orgl=getOCACHE()["%s_kuerzel"%orgtyp].get(kuerzel)
    if not orgl: return None
    orgl.sort(key=lambda x: x.gueltig_ab, reverse=True)
    ret=""
    for g in orgl:
        fkname=getMAName(objFromCache("mitarbeiter_qnr",getS(g,"q_nummer")))
        gab=getS(g, "gueltig_ab")
        if not gab:gab="1970-01-01"
        if gab and fkname:
            if ret:ret+="\n"
            ret+="%s %s"%(fkname,gab)
    return ret


def addIparamVals(iparam):
    iparam.domain="carit"
    iparam.dtnow=datetime.date.today()
    iparam.dtnowstr=iparam.dtnow.strftime(DFMT)
    if not iparam.adatumstr:iparam.adatumstr = iparam.dtnowstr
    if not iparam.adatum:
        iparam.adatum = iparam.dtnow
        iparam.ad_start = iparam.adatum.replace(day=1)
        iparam.eomday = eomday(iparam.adatum.year, iparam.adatum.month) 
        iparam.ad_end = iparam.adatum.replace(day=iparam.eomday)


def validateValue(iparam,cell,ret):
    if iparam and hasattr(iparam, "user") and iparam.user.startswith("SYS_"): return
    if iparam and iparam.otype=="sonderzahlungen" and cell.field in ["art_der_sonderzahlung"]:
        linf=getLinf(iparam)
        if linf.user =="nandini.sutrave" and cell.value!="Onboarding (ohne Erstattung)":
            ret.RetCode=u"Authorization failed (Code: WappHandler.2176->validateValue)"
        return

    if cell.field in ["prof_level","prof_level_new"] and iparam.otype != "planzak":
        g=getOCACHE()["gehalt"].get(cell.oid)
        car_path=getS(g,"car_path")
        if not car_path:
            ret.RetCode="Fehler: 'bitte zuerst den Career Track festlegen."
            return
        if cell.value and cell.value.startswith("P"):
            pl=int(cell.value[-1])
            if car_path in ["Expert", "Project"] and pl>5:
                ret.RetCode="Fehler: bei  diesem Career Track ist max. PL5 möglich."
                return
            if car_path in ["Leadership"] and pl<3:
                ret.RetCode="Fehler: bei  Leadership ist min. PL3 wählbar."
                return
        return
    if "rate_1_vom_blp"==cell.field:
        d=getOCACHE()["dienstfahrzeuge"].get(cell.oid)
        if getS(d,"dfzg_typ")!="Company Bike":
            ret.RetCode="Fehler: 'Rate' darf nur beim Company Bike editiert werden!"
        return
    if "datum" in cell.field or "geburtstag"  in cell.field or "gueltig_bis" in cell.field or cell.field in ["predialog_datum","uebernahme","gueltig_ab","rueckgabe","start","stop","ende","leasing_beginn","leasing_ende","ende_probezeit","geplanter_termin"]:
        if cell.value=="":return
        if cell.value=="unbefristet" and cell.field =="aufenthaltstitel_gueltig_bis":return
        d=getD(cell,"value")
        if not d or d.year<1970:
            trysplit=cell.value.split("-")
            if len(trysplit)<2:
                trysplit=cell.value.split(".")
            if len(trysplit)>=2:
                try:
                    if len(trysplit)==3 and not trysplit[2]:del(trysplit[2])
                    if len(trysplit)==3:
                        curyear=int(trysplit[2])
                        if curyear<100:curyear+=2000
                    else:
                        curyear=datetime.date.today().year
                    cell.value="%04ld-%02ld-%02ld"%(curyear,int(trysplit[1]),int(trysplit[0]))
                    d=getD(cell,"value")
                except:pass
        if not d:
            ret.RetCode="Fehler: '%s' ist kein gueltiges Datum!"%cell.value
            return
    if cell.field=="q_nummer":cell.value=cell.value.lower()
    if cell.field=="q_nummer" and (not cell.value.lower().startswith("q") and not cell.value.lower().startswith("x")): #(not cell.value.startswith("q") or cell.value.lower()!=cell.value):
        if cell.value:
            ret.RetCode=u"Fehler: '%s' ist keine gültige Q-Nummer!"%cell.value
            return
    if cell.field=="q_nummer" and iparam and iparam.otype in ["mitarbeiter"]:
        if iparam.otype+"_qnr" not in getOCACHE():
            getOCACHE()[iparam.otype+"_qnr"]=dictByPNum(iparam.otype)
        for otype in ["mitarbeiter"]:
            if cell.value in getOCACHE()[otype+"_qnr"]:
                mad=getOCACHE()[otype+"_qnr"]
                ma=mad.get(cell.value)
                if ma and ma.ext_user=="1":
                    setObjectAttribute(iparam,ma,"ext_user","0")
                    return
                ret.RetCode="Fehler: '%s' existiert bereits!"%cell.value
                return
        return
    dvalidate=None
    try:
        dvalidate=XEDB.DYNMODS[iparam.otype].validateValue
    except:
        pass
    if dvalidate:return dvalidate(iparam,cell,ret)

def implicitRoles(pnum,dtnowstr):
    #implicit roles
    rl=[]
    for orgt in ["abteilung","team"]:
        dl=dictByPNumList(orgt).get(pnum,[])
        for o in dl:
            if not o.gueltig_ab:continue
            gbis="9999-99-99" if not o.gueltig_bis else o.gueltig_bis
            if dtnowstr>=o.gueltig_ab and dtnowstr<=gbis:
                if orgt=="abteilung":rl.append("deptlead")
                if orgt=="team":rl.append("teamlead")
                break
    return rl


def loadRoles():#new version based on pnums instead of usernames
    XEDB.XSEROLES={}
    XEDB.USERROLES={}
    # explicit roles
    for ro in getObjects("roles"):
        if not ro.role:continue
        ass = []
        if ro.assigned:
            ass= loads(ro.assigned)
        for pnum in ass:
            getOrAdd(XEDB.USERROLES,pnum,[]).append(ro.role)
        XEDB.XSEROLES[ro.role] = ass

def handleSetValues(iparam,internal=0,nohistory=0):
    olist=[]
    ret=None
    if iparam.method=="SetValues":
        orgotype=subtype=None
        if hasattr(iparam,"otype") and iparam.otype[-1].isdigit() and not iparam.otype[-2].isdigit():# and not iparam.otype.endswith("2018")and not iparam.otype.endswith("360"):
            iparam.org_otype=iparam.otype
            iparam.otype=iparam.otype[:-1]
            orgotype=iparam.org_otype
        if iparam.otype in ["compbike"]:
            setfunc=getFunc(iparam,"set")
            setfunc(iparam)
        ret=CONT()
        ret.retd={}
        ret.RetCode="ok"
        #Nach SAP Migration keine Änderungen mehr:
        if iparam.otype in ["hr_krankheit","abwesenheit","absence"] and not iparam.sap_upload:
            aart=None
            if iparam.otype=="abwesenheit":
                cacheo=None
                if iparam.cells and iparam.cells[0].oid:
                    cacheo=getOCACHE()[iparam.otype].get(iparam.cells[0].oid)
                if cacheo:
                    aart=cacheo.art
                elif iparam.cells and iparam.cells[0].field=="art":
                    aart=iparam.cells[0].value
            if aart and aart not in ["Elternzeit", "Krank (> 6 Wochen)", "Teilzeit in Elternzeit"]:
                ret.RetCode="This not possible anymore (SAP migration)"
                return [],ret
        newids={}
        #Sonderfall Krankheit bei erstnaligem Eintrag:
        if iparam.otype in ["hr_krankheit"]:
            if iparam.cells and iparam.cells[0].oid.startswith("Rq"):
                ncell=CONT()
                ncell.oid=iparam.cells[0].oid
                ncell.field="q_nummer"
                ncell.value=iparam.cells[0].oid[1:]
                iparam.cells.append(ncell)

        for cell in iparam.cells:
            try:
                cell.value=cell.value.strip()
            except:pass
            if not internal:validateValue(iparam,cell,ret)
            if ret.RetCode!="ok":
                return olist,ret
            if cell.field==None:continue
            iparamuser=iparam.user
            linf=getLinf(iparam)
            if linf and linf.amuser and linf.amuser.real_q_nummer and iparam.otype!="trainings":
                iparamuser=linf.amuser.real_q_nummer
            if iparam.cells:
                if cell.oid.startswith("R"):
                    noid=newids.get(cell.oid)
                    if noid:cell.oid=noid
                if cell.oid.startswith("R"):
                    sql="INSERT INTO excelse_%s (xseuser,otype,val_string,oid,field) VALUES('%s','%s','%s','%s','%s') RETURNING tid;"%(XEDB.DOMAIN,iparamuser,iparam.otype,cell.value,cell.oid,cell.field,)
                    tid=runDBStatement(sql,1)
                    noid="I%09ld"%tid
                    newids[cell.oid]=noid
                    cell.oid=noid
                    sql="UPDATE excelse_%s set oid = '%s' where tid=%ld;"%(XEDB.DOMAIN,cell.oid,tid)
                    print(sql)
                    tid=runDBStatement(sql)
                    if subtype:
                        sql="INSERT INTO excelse_%s (xseuser,otype,val_string,oid,field) VALUES('%s','%s','%s','%s','%s');"%(XEDB.DOMAIN,iparamuser,iparam.otype,subtype,cell.oid,"__subtype",)
                        print(sql)
                        runDBStatement(sql)
                else:
                    try:
                        ivalue=cell.value.replace("'","''")
                    except:
                        ivalue=cell.value
                    if ivalue==None: ivalue =""
                    rd=None
                    if nohistory:
                        rd=genericSelect("SELECT tid from public.excelse_%s where otype='%s' and field='%s' and status is null and oid='%s'"%(XEDB.DOMAIN,iparam.otype,cell.field,cell.oid))
                        if rd:
                            sql="UPDATE excelse_%s SET xseuser='%s',val_string='%s' WHERE tid=%ld;"%(XEDB.DOMAIN,iparamuser,ivalue,rd[0].tid)
                    if not nohistory or not rd:
                        sql="INSERT INTO excelse_%s (xseuser,otype,val_string,oid,field) VALUES('%s','%s','%s','%s','%s');"%(XEDB.DOMAIN,iparamuser,iparam.otype,ivalue,cell.oid,cell.field,)
                    print("+++"+sql)
                    runDBStatement(sql)
                if orgotype:
                    de=getOrAdd(ret.retd,orgotype,CONT())
                else:
                    de=getOrAdd(ret.retd,iparam.otype,CONT())
                if not hasattr(de, "odict"): de.odict={}
                r=getOrAdd(de.odict,cell.oid,CONT())
                if iparam.otype not in ["urlaub"]: 
                    setattr(r, cell.field, cell.value)
                r.oid= cell.oid
                oc=objFromCache(iparam.otype,cell.oid,1)
                if getS(oc,"status")=="DELETED":
                    del getOCACHE()[iparam.otype][cell.oid]
                    oc=None
                    sql="UPDATE excelse_%s set status='DELETED' where oid = '%s';"%(XEDB.DOMAIN,cell.oid)
                    print(sql)
                    tid=runDBStatement(sql)
                if iparam.otype=="abwesenheit" and cell.field=="q_nummer":
                    r=getOrAdd(de.odict,cell.oid,CONT())
                if oc:setattr(oc, cell.field, cell.value)
                if oc and subtype and not hasattr(oc,"__subtype"):
                    oc.__subtype=subtype
                if oc and oc not in olist:
                    olist.append(oc)
        if orgotype:
            iparam.otype=orgotype
        
        #special case feedback form
        if iparam.otype=="fbpofo21":
            from shared_code.Services.getTalent21y import fixMissingFormData
            fixMissingFormData(iparam,[oc])

    return olist,ret

def addNLString(ins,add,sep="\n"):
    if ins==None:
        ins=""
    if ins:ins+=sep
    return ins+add

def addMonths(a, m):
    msum = a.month + m
    years = int((msum) / 12)
    months = msum % 12
    if months == 0:
        months = 12
        years -= 1
    return datetime.date(a.year + years, months, a.day)


def handleDeleteObject(iparam, force=0):
    """
    Handles the deletion of objects in the database.

    This function checks if the object type and conditions for deletion meet specific criteria before proceeding with the deletion.
    It supports a 'force' parameter to bypass certain checks, such as authorization for specific object types.

    Parameters:
    - iparam: An object containing parameters for the deletion, including the method, object type (otype), and object id (oid).
    - force (optional): A flag to force deletion without authorization checks. Defaults to 0 (false).

    Returns:
    - An instance of CONT with a RetCode indicating the result of the deletion attempt.
    """
    
    # Check if the method is specifically to delete an object
    if iparam.method == "DeleteObject":
        # Adjust the object type if the last character is a digit but the second to last is not
        if iparam.otype and iparam.otype[-1].isdigit() and not iparam.otype[-2].isdigit():
            iparam.org_otype = iparam.otype  # Preserve the original object type
            iparam.otype = iparam.otype[:-1]  # Modify the object type for processing
        
        # Check for authorization if not forced and object type is restricted
        if not force and iparam.otype in ["chargings"]:
            ret = CONT()
            ret.RetCode = "Not Authorized!"
            return ret
        
        #Nach SAP Migration keine Änderungen mehr:
        if iparam.otype in ["hr_krankheit","abwesenheit","absence"] and not iparam.sap_upload:
            aart=None
            if iparam.otype=="abwesenheit":
                cacheo=getOCACHE()[iparam.otype].get(iparam.oid)
                if cacheo:
                    aart=cacheo.art
            if not aart or aart  in ["Elternzeit", "Krank (> 6 Wochen)", "Teilzeit in Elternzeit"]:
                pass
            else:
                ret = CONT()
                ret.RetCode="This not possible anymore (SAP migration)"
                return ret

        # Attempt to check for conditions preventing deletion, if applicable
        try:
            check_delete = XEDB.DYNMODS[iparam.otype].checkDelete
            if check_delete:
                ret = CONT()
                check_delete(iparam, ret)
                if ret.RetCode and ret.RetCode != "ok":
                    return ret
        except:
            pass  # If an error occurs in the check, proceed without failing
        
        # Special case handling for a specific object type
        if iparam.otype == "grad360prov":
            iparam.otype = "grad360fb"
        
        # Insert a record to track the deletion
        sql = "INSERT INTO excelse_%s (xseuser,otype,val_string,oid,field,status) VALUES('%s','%s','%s','%s','%s','%s');" % (XEDB.DOMAIN, iparam.user, iparam.otype, "##DELETED##", iparam.oid, "oid", "DELETED")
        runDBStatement(sql)
        
        # Update the status of the object to 'DELETED'
        sql = "UPDATE excelse_%s set status = 'DELETED' where otype='%s' and oid='%s';" % (XEDB.DOMAIN, iparam.otype, iparam.oid)
        runDBStatement(sql)
        
        # Log the deletion
        logging.info("Deleting object...")
        logging.info(f"Otype: {iparam.otype}")
        logging.info(f"OID: {iparam.oid}")
        
        # Remove the object from the cache
        #Fix from IT4CARIT-1968: Reload Cache after IT4CARIT-1984
        del getOCACHE()[iparam.otype][iparam.oid]
        
        # Return a successful response
        ret = CONT()
        ret.RetCode = "ok"
        return ret

def handleFieldHistory(iparam):
    """
    Retrieves the history of changes made to a specific field of an object.

    This function looks up the change history for a given field of an object type specified in the input parameters.
    It supports handling special cases where the object type needs to be adjusted before querying the database.

    Parameters:
    - iparam: An object containing parameters for the query, including the method, object type (otype),
              and cells which contain object id (oid) and field information.

    Returns:
    - A result set containing the distinct history records for the specified field of the object,
      or None if the method is not 'GetFieldHistory' or no records are found.
    """
    
    # Initialize the return data to None
    rd = None
    
    # Check if the method is to get the field history
    if iparam.method == "GetFieldHistory":
        # Adjust object type for special cases
        if iparam.otype.startswith("bewerber"):
            otype2 = "bewerber"
        else:
            otype2 = iparam.otype
        
        # Further adjustment for specific object types
        if otype2 == 'pofo2018':
            otype2 = 'pofo'
        if otype2.startswith("pofomgmnt"):
            otype2 = "pofomgmnt"+iparam.pofoyear[2:]
        
        # Attempt to get the object dictionary for the adjusted object type
        tod = getObjectDict(otype2)
        t = tod.get(iparam.cells[0].oid)
        
        # Determine the correct object id to use for the history query
        if t is None:
            oid = objFromCache(otype2 + "_qnr", iparam.cells[0].oid).oid
        else:
            oid = t.oid
        
        # Perform the database query to fetch the field history
        rd = genericSelect("""
        SELECT distinct on (tid,field) tstamp,xseuser,val_string 
        FROM public.excelse_%s 
        where otype='%s' and field='%s' and status is null and oid='%s'
        order by tid desc,field""" % (XEDB.DOMAIN, otype2, iparam.cells[0].field, oid))
    
        # replace all q-numbers with real Names if possible
        for r in rd:
            maname=getMAName(r.xseuser)
            if maname and len(maname)>4:
                r.xseuser=maname
    # Return the query result set or None
    return rd

def isCont(o):
    return isinstance(o,CONT)

def getValueDef(o,default=None):
    if isCont(o):
        if default != None and o.value==None:
            return default
        return o.value
    if default != None and o==None:
        return default
    return o

def getValue(o):
    if isCont(o):
        return o.value
    return o

def mround(f):
    frac=f%1
    if frac >=0.5:
        return int(f)+1
    return int(f)

def rundenVolle(i,v):
    ret=int(mround(i/(v*1.0))*v)
    return ret

def trySet(m, mf, o, f=None):
    if not f: f = mf
    try:
        setattr(m, mf, getattr(o, f))
    except:
        setattr(m, mf, "#NV")


def dictMax(d,k,val):
    """collect the max value for a key in the dict"""
    if val==None:return
    v0=d.get(k)
    if v0==None:
        m=val
    else:
        m=max(v0,val)
    d[k]=m

def getUserByQNum(qnum):
    m=getOCACHE().get("mitarbeiter_qnr").get(qnum)
    if m:return m
    m=getOCACHE().get("planzak_qnr").get(qnum)
    if m:return m
    m=getOCACHE().get("planzak_qxnr").get(qnum)
    if m:return m
    m=getOCACHE().get("nomiko_qnr").get(qnum)
    if m:return m

def user2MA(uname):
    if uname.startswith("q") and uname[-1].isdigit():
        return getUserByQNum(uname)
    if uname:
        maun = getOrAdd(getOCACHE(), "mitarbeiter_uname", {})
        retl = maun.get(uname.lower())
    else:retl = None
    if not retl:
        ret = CONT()
        ret.q_nummer = "qUNKNOWN"
        return ret
    if len(retl) == 1:
        return retl[0]
    minad = ""
    for r in retl:
        ad = getS(r, "austrittsdatum")
        if not ad:
            retl = [r]
            return r
        if ad > minad:
            maxr = r
            minad = ad
    retl = [maxr]
    return maxr 

def getCarITEmail(m):
    ret = getS(m, "email_carit")
    if ret:return ret
    return getS(m, "email")

def isMyObject(ao,linf,otype=None):
    if otype=="pofofb":
        return linf.amuser.q_nummer==ao.recipient
    return getS(ao,"q_nummer")==linf.amuser.q_nummer

def getEmail2MA(iparam):
    email2ma = {}
    today = datetime.date.today().strftime("%Y-%m-%d")
    for matype in ["mitarbeiter", "planzak"]:
        for p in getOCACHE()[matype].values():
            if getS(p, "q_nummer") == "q510080":
                pass
            if matype == "planzak":
                beginn_anue_datum = getS(p, "beginn_anue_datum")
                ende_anue_datum = getS(p, "ende_anue_datum")
                if not beginn_anue_datum or not ende_anue_datum:continue
                if today < beginn_anue_datum or today > ende_anue_datum:continue
            elif not maAktiv(iparam, p, None, today):continue
            for email in ["email_bmw", "email_carit", "email"]:
                em = getS(p, email).lower()
                if em and em not in email2ma.keys():
                    email2ma[em] = p
    return email2ma

def XPY_Auth(iparam, linf,apptab_id=None):
    hrlead = "hrlead" in linf.roles 
    hreditor = "hreditor" in linf.roles 
    fuhrpark = "servicesFuhrpark"  in linf.roles 
    frontdesk = "servicesFrontdesk"  in linf.roles 
    external = "external"  in linf.roles
    hradmin = "hradmin"  in linf.roles
    ZAKNomiko= "ZAKNomiko" in linf.roles
    changeuser ="changeuser" in linf.roles

    TESTER_CHANGE_USER_DICT=os.environ["TESTER_CHANGE_USER_DICT"]

    if XEDB.DOMAIN == "carit" and iparam.method=="changeUser":
        logging.info("Searching TESTER_CHANGE_USER_DICT (%s) for %s"%(TESTER_CHANGE_USER_DICT,linf.amuser.sub))
        if TESTER_CHANGE_USER_DICT.find(linf.amuser.sub):
            if not "changeuser" in linf.roles:
                linf.roles.append("changeuser")
                logging.info("==>   changeUser appended")
            return 1

    if XEDB.DOMAIN == "carit" and iparam.otype in ["supportmgmnt","payarasupport","adpproblem"]:
        return 1

    if iparam.otype in ["sonderzahlungen"]:
        return "Sonderzahlungen" in linf.roles or hreditor

    if "Employee Data" in linf.roles:
        for svc in XEDB.APPLIST["employeedata"]["list"]:
            if svc["id"]==iparam.otype:
                return 1
    if "Health Management" in linf.roles:
        for svc in XEDB.APPLIST["healthmeasures"]["list"]:
            if svc["id"]==iparam.otype:
                return 1

    if XEDB.DOMAIN == "carit" and iparam.otype in ["nomiko"]:
        return "nomiko" in linf.roles 

    if XEDB.DOMAIN == "carit" and iparam.otype in ["planzak","nomiko","pz"]:
        return hrlead or ZAKNomiko

    if XEDB.DOMAIN == "carit" and iparam.otype in ["my_feedbacks","feedback_provider","encryption_data","encryption","feedback_results"]:
        hier=buildHierarchy(iparam)
        hm=hier.get(linf.amuser.q_nummer)
        return hm and hm.vertragsart in ["Befristet", "Unbefristet","ZAK","Expat","AG"]
    
    if XEDB.DOMAIN == "carit" and iparam.otype in ["targetmgmnt"]:
        hier=buildHierarchy(iparam)
        hm=hier.get(linf.amuser.q_nummer)
        return hm and hm.vertragsart in ["Befristet", "Unbefristet","Expat","AG"]


    if XEDB.DOMAIN == "carit" and iparam.otype in ["timemgmnt","time_logging","sap_absence","time_correction","team_calendar","dynamic_part_time","time_statement","project_calendar"]:
        hier=buildHierarchy(iparam)
        hm=hier.get(linf.amuser.q_nummer)
        return hm #(hm and hm.abteilung in ['JD-K']) or linf.user=='q295381'

    if XEDB.DOMAIN == "carit" and iparam.otype in ["daily_report"]:
        hier=buildHierarchy(iparam)
        hm=hier.get(linf.amuser.q_nummer)
        return (hm and hm.is_fk) #(hm and hm.abteilung in ['JD-K'] and hm.is_fk) or linf.user=='q295381'


    if XEDB.DOMAIN == "carit" and iparam.otype in ["worker"]:
        return linf.amuser.sub in ["qx60276","qxz11ba","qxz369h"]

    if XEDB.DOMAIN == "carit" and iparam.otype in ["all_fb_replies"]:
        return linf.amuser.real_q_nummer in ["qx60276","qxz11ba","qxz369h"] or linf.amuser.sub in ["qx60276","qxz11ba","qxz369h"]

    if XEDB.DOMAIN == "carit" and iparam.method =="Login":
        return linf.amuser.sub in ["tim.solutions"]
        
    if XEDB.DOMAIN == "carit" and iparam.otype in ["GetFINPersons"]:
        return linf.user=="tim.solutions"
    
    if frontdesk:
        return iparam.otype in ["dienstfahrzeuge"]
    
    if iparam.otype in ["autochange", "worker"]:
        return hradmin
    
    if XEDB.DOMAIN == "carit" and iparam.otype in ["roles"]:
        return intersect(linf.roles, ["rolesadmin"])
    
    if XEDB.DOMAIN == "carit" and not external and iparam.otype in ["employeess", "absence", "team_absence", "cal_permission", "standorthis22", "sonderzahlungen22", "chargings22", "chargingstat22", "adpsso22"]: 
        return 1

    if XEDB.DOMAIN == "carit" and not external and iparam.otype in ["adpsso"]:
        #Exceptions of Externals who are auth in ADP
        if linf.amuser.q_nummer in ["qx60276","qxz4j9d"]:return 1
        vertrag = getVertragByDate(linf.amuser.q_nummer, iparam.dtnowstr)
        return vertrag and getS(vertrag, "vertragsart") in ["Befristet", "Unbefristet", "Praktikant", "Werkstudent", "Expat", "Praktikant", "Werkstudent", "Studienabschlussarbeit", "Doktorand"]
    
    if XEDB.DOMAIN == "carit" and external and iparam.otype in ["orgaweb"]:
        return linf.roles != ["external"] or linf.amuser.email in ["manuel.muench@partner.bmw.de", "florian.fs.schneider@partner.bmw.de", "jonas.juerging@partner.bmw.de", "christian.cp.pelz@partner.bmw.de"]
    
    if 0 and XEDB.DOMAIN == "carit" and not external and iparam.otype in ["standorthis", "mitarbeiter"] and XEDB.HOSTNAME in ["PITSLG", "citbit-payarahr22"]:
        return 1
    
    if XEDB.DOMAIN == "carit" and iparam.otype in ["fb360results", "grad360fb", "grad360prov"]:
        return 1
    
    if XEDB.DOMAIN == "carit" and not external and iparam.otype in ["trainings"]:
        if 1 or XEDB.HOSTNAME in ["PITSLG", "citbit-payarahr"] or linf.amuser.q_nummer in XEDB.TRAININGS:
            return 1
    
    if XEDB.DOMAIN == "carit" and iparam.otype in ["chargings", "chargingstat", "skipped_chargings","charging_upload","meter_reading"]: #FORCE99
        if fuhrpark or hradmin:
            return 1
        hasemog = 0
        for d in getObjects("dienstfahrzeuge"):
            if getS(d, "q_nummer") != linf.amuser.q_nummer:
                continue
            if getS(d, "dfzg_typ") == "DienstFzg EmoG":
                hasemog = 1
                break
        return hasemog

    if iparam.method == "DeleteObject":
        if "hreditor" in linf.roles:return 1
        otype=getOtypeByOID(iparam.oid)
        iparam.otype=otype
        if fuhrpark and otype=="dienstfahrzeuge":
            return 1
        return 0

    if XEDB.DOMAIN == "carit" and iparam.otype in ["upload"]:
        return intersect(linf.roles, ["hradmin","hreditor"])

    if iparam.otype in ["review"]:
        return 1 if intersect(linf.roles, ["hradmin", "hrlead"]) else 0
    
    if XEDB.DOMAIN == "carit"  and iparam.otype in ["absence_overview"]:
        return 1 if intersect(linf.roles, ["teamlead", "deptlead"]) else 0

    if iparam.otype in ["changerequest", "form_data"]:
        return 1
    
    if iparam.method == "ValidateTables":
        return 1
    
    if iparam.otype in ["pofomgmngt"]:
        pass

    if iparam.otype in ["pofofb1"]:
        return 1
    
    if iparam.otype in ["pofofb2"]:
        mad = getOCACHE()["mitarbeiter_qnr"]
        selfqnr=linf.amuser.q_nummer
        ma=mad.get(selfqnr)
        eintr=getS(ma,"eintrittsdatum")
        if eintr:
            return 1
        return 0
    
    if iparam.otype in ["talent21y"]:
        return hrlead
    
    if iparam.otype in ["fkmitarbeiter","pofomgmnt","pofomgmnt2","talentgrid","talentperffit","pofo_consolidation"]:
        hastalentauth = "teamlead" in linf.roles or "deptlead" in linf.roles or "hrlead" in linf.roles
        return hastalentauth

    if iparam.otype in ["fb360results", "fb360results2"]:
        ma = user2MA(linf.user)
        if ma and getS(ma, "email_bmw") in ["christine.regler@bmw.de", "marina.sa.stierstorfer@bmw.de", "elena.buettner@bmw.de", "marco.steigenberger@bmw.de", "ludovic.martorell@bmw.de", "christian.manz@bmw.de", "Frank.Bieler@bmw.de", "alex.reif@bmw.de", "Michael.Boettrich@bmw.de", "christian.enchelmaier@bmw.de", "andreas.klenk@bmw.de", "timo.endres@bmw.de", "frank.jene@bmw.de", "andrea.wettke@bmw.de", "steffen.weiss@bmw.de", "andreas.falk@bmw.de", "Daniel.Nettesheim@bmw.de", "michael.chambers@bmw.de", "Kai-uwe.Balszuweit@bmw.de", "holger.grandy@bmw.de", "oliver.simeth@bmw.de", "matthias.behr@bmw.de", "fergal.lynch@bmw.de", "mariana.isac@bmw.de", "jochen.boehm@bmw.de", "wira.tirta@bmw.de", "Norbert.Fias@bmw.de", "sheng.chang@bmw.de", "ajhan.kopov@bmw.de", "tillmann.schumm@bmw.de", "sukhbansbir.boparai@bmw.de", "serhiy.smirnov@bmw.de", "girish.subramanian@bmw.de"]:
            return 1
        return 0

    if XEDB.DOMAIN == "carit" and not external and iparam.otype in ["GetPersonalInfo", "orgaweb", "urlaubmss"]:
        return 1

    if iparam.method == "SetValues" and "readonly" in linf.roles:
        if "servicesElidaNachweis" in linf.roles:
            if len(iparam.cells) == 1 and iparam.cells[0].field == "elida_nachweis":
                return 1
        return 0
    
    if "hreditor" in linf.roles:return 1
    for serv in linf.roles:
        if serv.startswith("services"):
            ls = XEDB.XSESERVICES[serv]
            if iparam.otype in ls:
                return 1
    return 0

def cleanOrgaRels(iparam):
    for orgtype in ["team", "abtlg"]:
        teaml = getOCACHE().get("qnr2" + orgtype)
        if not teaml:return
        torem1 = []
        for qnr, tl in teaml.items():
            if qnr == "q339339":
                pass
            torem2 = []
            for t in tl:
                gab = getS(t, "gueltig_ab", "0000-00-00", 1)
                gbis = getS(t, "gueltig_bis", "9999-99-99", 1)
                if not gab or iparam.dtnowstr < gab or iparam.dtnowstr > gbis:
                    torem2.append(t)
                    continue
            for rem in torem2:
                tl.remove(rem) 
            if not tl:
                torem1.append(qnr)
        for rem in torem1:del teaml[rem] 

def addGehaltByQNR(oc, qnr):
    otypepnr = "gehalt_qnr" if XEDB.DOMAIN == "carit" else "vgehalt_vnr" 
    gld = getOrAdd(getOCACHE(), otypepnr, {})
    gl = getOrAdd(gld, qnr, [])
    if oc not in gl:
        gl.append(oc)
        if not hasattr(oc, "gueltig_ab"): oc.gueltig_ab = ""    
        gl.sort(key=lambda x: x.gueltig_ab, reverse=True)

def changeUser(iparam):
    li = getLinf(iparam)
    ret = CONT()
    hrlead = "hrlead" in li.roles
    if not intersect(li.roles, ["changeuser"]):
        logging.info("%s did not match changeuser" % li.roles)
        ret.RetCode = "You can not change user. If the role was recently assigned to you, please log out and log in again."
        return ret
    if li.amuser.sub=="q509229" and iparam.fparams.user =="q156581":
        ret.RetCode = "You can not change the user to Christine Regler"
        return ret

    li.user = iparam.fparams.user 
    li.roles = ["changeuser"]
    li.newchange = 1
    if not hasattr(li.amuser, "real_q_nummer"):
        li.amuser.real_q_nummer = li.amuser.sub

    for motype in ["mitarbeiter", "planzak", "nomiko"]:
        m = objFromCache("%s_qnr" % motype, iparam.fparams.user)
        if m:
            break

    li.amuser.q_nummer = iparam.fparams.user

    q_nummer=getS(m,"q_nummer")
    qx_nummer=getS(m,"qx_nummer")

    if qx_nummer:
        li.amuser.qx_nummer = qx_nummer.lower()
    if qx_nummer and not q_nummer:
        li.amuser.q_nummer = li.amuser.qx_nummer

    li.user = li.amuser.q_nummer
    li.amuser.name = getMAName(m)
    if 1:
        ret.retd = CONT()
        ret.retd.name = getMAName(m)
        ret.retd.qnum = "[TEST]"
    return ret

def handleSpecialMethods(iparam,linf):
    """
    Handle special methods based on the 'iparam.method' value.

    This function processes different methods specified in 'iparam.method' and performs the corresponding operations. 
    Supported methods include:
    - "GetUserInfo": Retrieves user information for logged in user
    - "GetAppList": Retrieves the list of applications for Payara HR Starting Page
    - "GetMAList": Retrieves a list of employees that we use for MAPicker in frontend (usually for cstate.malist in frontend)
    - "changeUser": Changes user information for testing purposes by another function (changeUser)

    Parameters:
    iparam (object): Contains parameters needed for method execution. (otype, method etc.)
    linf (object): Contains user and session information.

    Returns:
    object: The result of the method execution, usually a CONT() object with relevant data.
    """

    implemented_methods = ["GetUserInfo",'GetAppList',"GetMAList",'changeUser']

    if iparam.method not in implemented_methods:
        return
    
    ret=CONT()

    if iparam.method == "changeUser":
        return changeUser(iparam)

    if iparam.method =="GetUserInfo":
        ret.RetCode = "ok"
        if not linf.amuser.q_nummer:linf.amuser.q_nummer=linf.amuser.sub
        ret.retd = {'name': linf.amuser.name, 'qnum': linf.amuser.q_nummer, 'user_num': linf.amuser.q_nummer}
        logging.info("### 2817 GetUserInfo %s (%s)" % (linf.amuser.name, getS(linf.amuser, "q_nummer")))
        return ret
    
    if iparam.method == "GetAppList":
        if XEDB.DOMAIN == "carit":
            getOCACHE()["mitarbeiter_qnr"]=dictByPNum("mitarbeiter")
        cleanOrgaRels(iparam)
        iparam.user = linf.user
        retl = XEDB.APPLIST
        ret = {}
        hastalentauth = "teamlead" in linf.roles or "deptlead" in linf.roles or "hrlead" in linf.roles
        for attr, value in retl.items():
            if attr in ["managerss"]:
                if hastalentauth: 
                    ret[attr] = value
                    if "hreditor" not in linf.roles:
                        newvlist = []
                        for lm in value['list']:
                            if lm['id'] != "gehband":
                                newvlist.append(lm)
                        value['list'] = newvlist
                        ret[attr] = value
            elif attr == "admin":
                if linf.user=="ext.stefan.spiess":
                    if "hradmin" not in linf.roles:
                        linf.roles.append("hradmin")
                if "hradmin" in linf.roles: 
                    ret[attr] = value
            elif 'list' in value:
                authorizedAppsList = []
                for appTab in value['list']:
                    if appTab['id'] == "talentletter":
                        if "hreditor" in linf.roles:
                            authorizedAppsList.append(appTab)
                        continue
                    iparam.cells = []
                    iparam.otype = appTab['id']
                    iparam.method = "GetValues"
                    getValueAuthorization = XPY_Auth(iparam, linf,appTab['id'])
                    iparam.method = "SetValues"
                    setValueAuthorization = XPY_Auth(iparam, linf,appTab['id'])
                    if getValueAuthorization or setValueAuthorization:
                        authorizedAppsList.append(appTab)
                    
                if len(authorizedAppsList) > 0:
                    ret[attr] = {'name': value['name'], 'list': authorizedAppsList}
            elif attr in ['excel_payroll', 'excel_reports', 'series_email']:
                if "hreditor" in linf.roles:
                    ret[attr] = value
            else:
                if not intersect(["servicesFrontdesk", "external"], linf.roles):
                    ret[attr] = value

            ma=MA_DICT().get(linf.amuser.q_nummer)
            if ma:
                ldreminder=getD(ma,"last_datareminder")
                if ldreminder:
                    remind_gap= (iparam.dtnow-ldreminder)
                if getS(ma,"oid") and not ldreminder or remind_gap.days>=182 and ma.ext_user!="1":
                    ret["datareminder"]=1
                    setObjectAttribute(iparam, ma, "last_datareminder", iparam.dtnowstr)

        return ret
    
    if iparam.method == "GetMAList":
        if XEDB.DOMAIN == "carit":
            ret = CONT()
            linf = getLinf(iparam)
            ma = user2MA(linf.user)
            ret.userinf = CONT()
            ret.userinf.q_nummer = ma.q_nummer
            ret.userinf.name = getMAName(ma.q_nummer)
            mall = getOCACHE()["mitarbeiter"].values()
            iparam.show_expired = ""
            planzaks = getService("planzak", iparam).values()
            ret.malist = []

            #Add employees to malist
            for m in mall:
                edd = getS(m, "eintrittsdatum")
                add = getS(m, "austrittsdatum")
                ext_permission=getN(m,"ext_permission")
                
                if not ext_permission:
                    if add and add<iparam.dtnowstr:continue
                    if not edd:continue
                else:
                    if add and add<iparam.dtnowstr:
                        continue

                mm = CONT()
                mm.text = getMAName(m)
                mm.value = m.q_nummer
                ret.malist.append(mm)

            #Add planzaks to malist
            for m in planzaks:
                if not getS(m, "q_nummer"):continue
                mm = CONT()
                mm.text = getMAName(m)
                mm.value = m.q_nummer
                ret.malist.append(mm)

            ret.malist.sort(key=lambda x: x.text, reverse=False)
            mm = CONT()
            mm.text = ""
            mm.value = ""
            ret.malist.insert(0, mm)
            return ret
        else:
            pass
            return ret

def getUserDict(iparam,chkdate=None,onlyactive=1):
    if not chkdate:
        chkdate=iparam.dtnowstr if not iparam.eval_date else iparam.eval_date[:10]
    mad=dictByPNum("mitarbeiter")
    for matype in ["planzak","nomiko"]:
        for o in getObjects(matype):
            if not o.q_nummer and not o.qx_nummer:
                continue
            if not o.beginn_anue_datum:
                continue
            if o.beginn_anue_datum>chkdate:
                continue

            m=None
            if o.q_nummer:
                m=mad.get(o.q_nummer)
            if not m and o.qx_nummer:
                m=mad.get(o.qx_nummer)
            if m:
                mbeginn=m.beginn_anue_datum if m.beginn_anue_datum else m.eintrittsdatum
                if mbeginn:
                    if o.beginn_anue_datum<mbeginn:continue
            if o.q_nummer:mad[o.q_nummer]=o
            if o.qx_nummer:mad[o.qx_nummer]=o
    getOCACHE()["mitarbeiter_qnr"]=mad
    return mad


def getOrgByDateOLD(ma,orgtyp,kuerzel,dstring1):
    """TODO: Add documentation.

    Returns: g (any)
    """
    if ma:
        austritt=getS(ma,"austrittsdatum","9999-99-99",True) #True means that empty string has to be replaced as if not existent
        dstring=min(dstring1,austritt)
    else:
        dstring=dstring1
    #dummy=getOCACHE()["%s_kuerzel"%orgtyp]
    if not kuerzel:
        return
    if 0 and "-" in kuerzel and kuerzel.endswith("DR"):
        kuerzel=kuerzel[:-2]
        orgtyp="abtlg"
    orgl=getOCACHE()["%s_kuerzel"%orgtyp].get(kuerzel)
    if not orgl: return None
    try:
        orgl.sort(key=lambda x: x.gueltig_ab, reverse=True)
    except:
        for g in orgl:
            if getS(g, "gueltig_ab","")=="":
                g.gueltig_ab="1970-01-01"
        orgl.sort(key=lambda x: x.gueltig_ab, reverse=True)
        
    for g in orgl:
        if hasattr(g, "gueltig_ab"):
            if dstring>=g.gueltig_ab:
                break
    if g.kuerzel in ["NoTeam","#NV","NoAbtlg"]:
        return None
    return g


class HIERA():
    def __init__(self,iparam,CHKDATE=None,onlyactive=1) -> None:
        self.hiera=buildHierarchy(iparam,CHKDATE,onlyactive)
        self.orgdict=self.hiera["ORGDICT"]
    def get(self,qnum):
        return self.hiera.get(qnum)
    def getTeamlead(self,qnum):
        rhier=self.hiera.get(qnum)
        if rhier and rhier.team:
            team=self.orgdict.get(rhier.team)
            if team:
                return team.q_nummer
    def getAbteilungslead(self,qnum):
        rhier=self.hiera.get(qnum)
        if rhier and rhier.abteilung:
            abteilung=self.orgdict.get(rhier.abteilung)
            if abteilung:
                return abteilung.q_nummer

def buildHierarchy(iparam,CHKDATE=None,onlyactive=1):
    HIERA={}
    if not CHKDATE:CHKDATE= iparam.dtnowstr
    gfuehrer=getGFuehrer(CHKDATE)
    fkdict={}
    orgdict={}
    if 1 or not mall:
        mad=getUserDict(iparam, CHKDATE)
        mall={}
        for mqnum,m in mad.items():
            if onlyactive:
                if not maAktiv(iparam,m,adatumstr=CHKDATE):continue
            nm=CONT()
            setFValue(nm,m,"q_nummer")
            if m.beginn_anue_datum:
                v=CONT()
                setFValue(v,m,"team")
                setFValue(v,m,"abteilung")
                nm.vertragsart="ZAK"
            else:
                v=getVertragByDate(nm.q_nummer,CHKDATE)
                setFValue(nm,v,"vertragsart")
            setFValue(nm,v,"team")
            if nm.team and nm.team not in orgdict:
                orgentry=getOrgByDateOLD(nm,"team",nm.team,CHKDATE)
                if orgentry and orgentry.gueltig_bis and orgentry.gueltig_bis < CHKDATE:
                    pass
                else:
                    orgdict[nm.team]=orgentry
                    if orgdict[nm.team]:
                        fkdict[orgdict[nm.team].q_nummer]=("team",nm.team)
            setFValue(nm,v,"abteilung")
            if nm.abteilung=="JC-3":
                print
            if nm.abteilung and nm.abteilung not in orgdict:
                orgentry=getOrgByDateOLD(nm,"abtlg",nm.abteilung,CHKDATE)
                if orgentry and orgentry.gueltig_bis and orgentry.gueltig_bis < CHKDATE:
                    pass
                else:
                    orgdict[nm.abteilung]=orgentry
                    if orgdict[nm.abteilung]:
                        fkdict[orgdict[nm.abteilung].q_nummer]=("abteilung",nm.abteilung)
                    orgdict[nm.abteilung+"DR"]=getOrgByDateOLD(nm,"abtlg",nm.abteilung,CHKDATE)
            setFValue(nm,v,"hauptabteilung")
            if nm.hauptabteilung and nm.hauptabteilung not in orgdict:
                orgdict[nm.hauptabteilung]=getOrgByDateOLD(nm,"hauptabteilung",nm.hauptabteilung,CHKDATE)
                if orgdict[nm.hauptabteilung]:
                    fkdict[orgdict[nm.hauptabteilung].q_nummer]=("hauptabteilung",nm.hauptabteilung)
                orgdict[nm.hauptabteilung+"DR"]=getOrgByDateOLD(nm,"hauptabteilung",nm.hauptabteilung,CHKDATE)
            mall[nm.q_nummer]=nm
            if m.qx_nummer:
                nm.qx_nummer=m.qx_nummer
    gfuehrer=getGFuehrer(CHKDATE)
    orgdict["gmbh"]=gfuehrer
    fkdict[gfuehrer.q_nummer]=("gmbh","gmbh")
    for m in mall.values():
        if m.q_nummer=='q295381':
            print
        fk=fkdict.get(m.q_nummer)
        if fk:
            m.is_fk=1
            if fk[0]=="team":
                m.iam_tl=1
            if fk[0]=="abteilung": m.iam_al=1
            if fk[0]=="hauptabteilung": m.iam_hal=1
            if fk[0]=="gmbh": m.iam_gf=1
        HIERA[m.q_nummer]=m
        if m.qx_nummer:
            HIERA[m.qx_nummer]=m
        if not m.is_fk and m.team and m.team not in ["NoTeam","#NV"]:
            if orgdict.get(m.team):
                m.myboss=orgdict[m.team].q_nummer
        if  m.team:
            getOrAdd(HIERA,m.team,[]).append(m.q_nummer)
        if m.iam_tl and m.abteilung:
            if not m.abteilung in orgdict or not orgdict[m.abteilung]:
                logging.info("####HIERARCHY PROBLEM, no entry for %s"%dumps(m))
                m.myboss=None
            else:
                m.myboss=orgdict[m.abteilung].q_nummer
        if  m.abteilung and (m.is_tl or (m.team and "-" in m.team and m.team.endswith("DR"))):
            getOrAdd(HIERA,m.abteilung,[]).append(m.q_nummer)

        if m.iam_al and m.hauptabteilung and m.hauptabteilung in orgdict and orgdict[m.hauptabteilung]:
            m.myboss=orgdict[m.hauptabteilung].q_nummer
        if  m.hauptabteilung and (m.is_al):
            getOrAdd(HIERA,m.hauptabteilung,[]).append(m.q_nummer)
        if m.hauptabteilung and m.iam_hal:
            m.myboss=gfuehrer.q_nummer
            getOrAdd(HIERA,"gmbh",[]).append(m.q_nummer)

        
    HIERA["ORGDICT"]=orgdict
    return HIERA

def getHierarchyUNUSED(iparam,qnum,CHKDATE=None):
    if not CHKDATE:CHKDATE=iparam.dtnowstr
    v=getVertragByDate(qnum,CHKDATE)
    ret=CONT()
    if v.team:
        ret.team=v.team
        getOrgByDate()
    if v.abteilung:
        ret.abteilung=v.abteilung


class CRET(CONT):
    def __init__(self, rc):
        self.RetCode = rc

def getAttributes(iparam, r=None):

    getFunc(iparam)

    smodule=XEDB.DYNMODS[iparam.otype]

    attribs=[]
    if hasattr(smodule,"getUIGRID"):
        try:
            gridinfo=smodule.getUIGRID(iparam,internal=1)
        except:
            gridinfo=smodule.getUIGRID(iparam)
    else:
        if hasattr(smodule,"UIGRID"):
            gridinfo=smodule.UIGRID

    if gridinfo:
        for uif in gridinfo["fields"]:
            attribs.append(uif["name"])

    return attribs

def getBAVNumbers(iparam,qnum):
    uparam=dictByAttr("urlaub_p","jahr")
    ret={}
    szbyqnum=dictByPNumList("sonderzahlungen")
    mysz=szbyqnum.get(qnum,[])
    for year in range(2022,iparam.dtnow.year+1):
        r=CONT()
        r.year=year
        r.q_nummer=qnum
        r.extra_conversion=r.base_amount=r.per_month_ag=r.per_month_an=0
        
        yearstr=str(year)
        for month in range(1,13):
            chkdate="%ld-%02ld-10"%(year,month)
            mpbav=MAPBAV(chkdate)
            g=getGehaltByDate(qnum,chkdate)
            bavd=mpbav.map(g.prof_level)
            r.base_amount+=getN(bavd,"betrag")
            if g.bav_conv_activated in ["YES","1"]:
                r.per_month_ag+=getN(bavd,"wandlung")
                r.per_month_an+=getN(bavd,"wandlung")
        for sz in mysz:
            if sz.art_der_sonderzahlung=='BAV Entgeltumwandlung' and sz.zahlungsdatum and sz.zahlungsdatum.startswith(yearstr):
                r.extra_conversion-=getN(sz,"sonderzahlung")
        r.sum=r.extra_conversion+r.base_amount+r.per_month_ag+r.per_month_an
        basev=int(uparam[str(year)].bemessungsgrenze_rvav.replace(",00","").replace(".",""))
        r.soz_vers_und_steurfrei= basev * 0.04*12
        r.nur_steuerfrei = basev * 0.08*12


        ret[year]=r
    return ret
    #getOCACHE()["gehalt_qnr"].get(qnum)


def handleGenHTML(iparam):
    refreshDicts()

    linf = getLinf(iparam)
    if "hreditor" not in linf.roles:
        iparam.args = None
    qnum=linf.amuser.q_nummer
    uma=getOCACHE()["mitarbeiter_qnr"].get(qnum)
    if not uma: 
        return CRET("No employee found: " + iparam.user)
    m = uma
    ret = u"""<head>  <meta charset="cp1252"> <title>Personal Info</title> 
    <style> @media print { h7 {page-break-after:always;} } </style>  
<style>
    * {
      font-family: sans-serif;
    }

table, th, td {
    border: 1px solid black;
    border-collapse: collapse;
    white-space:pre-wrap; word-break:break-word;
}
th {
  word-break:break-all;
  white-space:normal;
  word-wrap:break-word;
  BACKGROUND-COLOR: #bbf0b6;
}

hr {
    display: block;
    height: 1px;
    border: 0;
    border-top: 1px solid;
    margin: 1em 0;
    padding: 0;
}</style>    </head> """
    iparam.otype = "mitarbeiter"
    ret += u"<h2>Downloads:<h2>"   
    butext='<button style="BACKGROUND-COLOR: green; height: 50px;color: white; font-weight: bold"  onclick="'+"window.open('https://myelida-p.bmwgroup.net/myElida.html#!/home?sprache=en', '_blank').focus();"+'">Portfolio- and salary letters (ELIDA)</button><br><br>'
    ret += butext

    ret += u"<h2>Stammdaten<h2>" + handleHTML(iparam, [m], 1,attribs=["q_nummer","nachname","vorname","weitere_vornamen","geschlecht","titel","nationalitaet","geburtsname","geburtsdatum","behinderungsgrad","geburtsort","konfession","username","email","email_carit","iban","bic","sozialversicherungsnr","steuernummer","status_vertrag_nur_fuer_studen","ansprechpartner_student","eintrittsdatum","ende_probezeit","konzerneintrittsdatum","austrittsdatum","emergency_contact","notfall_mobil","notfall_festnetz","private_email","private_phone","private_landline"],otype="mitarbeiter")    
    
    if 1 or XEDB.HOSTNAME in ["PITSLG", "citbit-payarahr"]:
        extra = CONT()
        for attr in ['emergency_contact','notfall_mobil','notfall_festnetz',"private_email","private_phone","private_landline"]:#  ["q23_notfall_festnetz", "q22_notfall_mobil", "q24_emergency_contact", "privateEmail", "privateTelefonnummer"]:
            setattr(extra, attr, getS(uma, attr))
        ret += u'<br><button style="BACKGROUND-COLOR: green; height: 50px;color: white; font-weight: bold" onclick="getByJForm(\'Emergency_contact.html\',null,\'%s\',true);">Enter Emergency Contact</button><br><br>' % object2bas64(extra) #base64.b64encode(dumps(extra))
        ret += u'<button style="BACKGROUND-COLOR: green; height: 50px;color: white; font-weight: bold" onclick="getByJForm(\'Private_Data.html\',null,\'%s\',true);">Enter Private Data</button><br>' %  object2bas64(extra) #base64.b64encode(dumps(extra))

        kidsqnr=dictByPNum("children").get(qnum)
        if kidsqnr:
            extra=bas642object(kidsqnr.formvals)
            del extra.validatedNewRequiredFieldIDs
            del extra.embedUrl
            kids=loads(kidsqnr.kids)
            nlite=100
            for k in kids:
                setattr(extra,"lite_mode_%ld"%nlite,k)
                nlite+=1

        ret += u'<br><button style="BACKGROUND-COLOR: green; height: 50px;color: white; font-weight: bold" onclick="getByJForm(\'Freiwillige_Selbstauskunft_gegenüber_dem_Arbeitgeber_zur_Anzahl_der_berücksichtigungsfähigen_Kinder_zur_Ermittlung_des_Beitrages_zur_sozialen_Pflegeversicherung_nach_§_55_Abs__3_SGB_XI.html\',null,\'%s\',true);">Enter self-report number of children</button><br>' % object2bas64(extra) #base64.b64encode(dumps(extra))

        if kidsqnr:
            bdates=loads(kidsqnr.kids)
            if bdates:
                bds=""
                for bd in bdates:
                    if bds:bds+=", "
                    bds+=bd
                ret+=" - Birthdays reported: "+bds
            ret+="<br>"
    import shared_code.Services.getAdpdiff
    reload(shared_code.Services.getAdpdiff)
    chkdate=datetime.date(iparam.dtnow.year,iparam.dtnow.month,eomday(iparam.dtnow.year,iparam.dtnow.month))
    chkdate=chkdate+timedelta(days=1)
    chkdate=date2String(chkdate)
    chkdate=max(chkdate,"2024-01-01")
    geh_in_month=shared_code.Services.getAdpdiff.getGehaltInMonth(iparam,chkdate[:7],qnum)
    iparam.otype = "gehalt"

    SWITCH_OFF_GEHALT=1

    if XEDB.ENVIRONMENT in ["TEST","INT"]:
        if linf.amuser and linf.amuser.real_q_nummer=="qx60276":
            SWITCH_OFF_GEHALT=0

    if SWITCH_OFF_GEHALT:
        ret+="<hr/><h7></h7>\n<h2>The salary history is temporarily switched off due to the ongoing salary planning process.<h2>"    
    else:
        ret += "<hr/><h7></h7>\n<h2>Gehaltshistorie<h2>" + handleHTML(iparam, getOCACHE()["gehalt_qnr"].get(qnum), 0, 0,attribs=["q_nummer","gueltig_ab","wochenstunden","prof_level","car_path","aktuelle_einstufung","monats_gehalt_vollzeit","praemien_grundwert_vollzeit","direktv_neu","zweit_wohnung","vwl","bav_conv_activated","bav_conv"],otype="gehalt")
    extra = CONT()
    extra.name = getMAName(uma)
    extra.__textreplace = CONT()
    #chkdate=iparam.dtnow_str#  "2024-01-01"
    extra.__textreplace.PHR_name = getMAName(uma)
    extra.__textreplace.PHR_team = getTeam(qnum,iparam.dtnowstr)
    mpbav=MAPBAV(chkdate)
    cgehalt=getGehaltByDate(qnum,chkdate)
    bavd=mpbav.map(cgehalt.prof_level)
    extra.__textreplace.PHR_level = cgehalt.prof_level
    extra.__textreplace.PHR_bavbase=bavd.betrag
    extra.__textreplace.PHR_optional=bavd.wandlung
    extra.__textreplace.PHR_today=chkdate
    extra.__textreplace.PHR_thisyear=chkdate[:4]
    extra.__textreplace.PHR_yearmonth=chkdate[:7]
    extra.__textreplace.PHR_sumpaid= getN(geh_in_month,"direktv_neu") if geh_in_month.direktv_neu else 0
    extra.__textreplace.PHR_sumpaid+= int(bavd.betrag) + 2*int(bavd.wandlung) if cgehalt.bav_conv_activated in ["YES","1"]  else int(bavd.betrag)
    bav_numbers=getBAVNumbers(iparam,qnum)
    extra.__textreplace.PHR_sozfrei= bav_numbers[int(chkdate[:4])].soz_vers_und_steurfrei#7550 * 0.04 if getN(cgehalt,"monats_gehalt_vollzeit")>7550 else getN(cgehalt,"monats_gehalt_vollzeit")*0.04
    extra.__textreplace.PHR_steuerfrei = bav_numbers[int(chkdate[:4])].nur_steuerfrei#7550 * 0.08 if getN(cgehalt,"monats_gehalt_vollzeit")>7550 else getN(cgehalt,"monats_gehalt_vollzeit")*0.08
    extra.__textreplace.PHR_stfrei_monthly=mround(extra.__textreplace.PHR_steuerfrei /12)
    extra.__textreplace.PHR_szfrei_monthly=mround(extra.__textreplace.PHR_sozfrei /12)
    extra.__textreplace.PHR_chosenbav="already chosen a monthly conversion" if cgehalt.bav_conv_activated in ["YES","1"] else "NOT chosen a monthly conversion"
    extra.__textreplace.PHR_actordeact="De-activate" if cgehalt.bav_conv_activated in ["YES","1"] else "Activate"
        
    extra.q_nummer = qnum

    iparam.otype=None
    first=1
    #bav_numbers_values=list(bav_numbers.values())
    for b in bav_numbers.values():
        b.year=str(b.year)
        if b.year==extra.__textreplace.PHR_thisyear:
            b.year+=" (forecast)"
            extra.__textreplace.PHR_bavsum=b.sum
            if b.extra_conversion:
                extra.input_251=b.extra_conversion
                extra.__textreplace.PHR_bavsum-=b.extra_conversion
        if b.sum>b.soz_vers_und_steurfrei:
            b.soz_vers_und_steurfrei= '<span style="background-color:#f2e52e;">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;%ld&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span>'%int(b.soz_vers_und_steurfrei)
        if b.sum>b.nur_steuerfrei:
            b.nur_steuerfrei= '<span style="background-color:#f2e52e;">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;%ld&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span>'%int(b.nur_steuerfrei)

    extra.input_250=[]
    for monthi in range(iparam.dtnow.month,iparam.dtnow.month+12):
        dropdyear=iparam.dtnow.year+1 if monthi>11 else iparam.dtnow.year
        monthir=(monthi%12)+1
        extra.input_250.append("%s %ld"%(num2monthEng(monthir),dropdyear))
    extra_styles={"soz_vers_und_steurfrei": 'style="background-color:#a5a8a6;"',"nur_steuerfrei": 'style="background-color:#a5a8a6;"'}
    #ret += "<hr/><h7></h7>\n<h2>BAV - history</h2><b>IMPORTANT : </b> all EUR amounts are cumulated for the respective year and the current year is a forecast" + handleHTML(iparam, bav_numbers.values(), 0, 0,extra_styles=extra_styles,attribs_trans={"q_nummer":0,"year":0,"base_amount":"bav grundbetrag","per_month_ag":"entgeltumwandlung ag","per_month_an":"entgeltumwandlung an","extra_conversion":"jährliche umwandlung","sum":"gesamt", "soz_vers_und_steurfrei":"st und sv frei","nur_steuerfrei":"st frei"})# otype="bav_numbers")    
    ret += "<hr/><h7></h7>\n<h2>BAV - history</h2><b>" + handleHTML(iparam, bav_numbers.values(), 0, 0,extra_styles=extra_styles,attribs_trans={"q_nummer":0,"year":"jahr","base_amount":"bav grundbetrag","per_month_ag":"entgeltumwandlung ag","per_month_an":"entgeltumwandlung an","extra_conversion":"jährliche umwandlung","sum":"gesamt", "soz_vers_und_steurfrei":"st und sv frei","nur_steuerfrei":"st frei"})# otype="bav_numbers")    
    ret += u'<button style="BACKGROUND-COLOR: green; height: 50px;color: white; font-weight: bold" onclick="getByJForm(\'BAV_Status.html\',null,\'%s\',true);">Edit your monthly conversion</button>&nbsp;&nbsp;&nbsp;&nbsp;' %  object2bas64(extra)
    dummy=iparam.dtnowstr[5:]
    if (XEDB.ENVIRONMENT in ["TEST","INT"] and extra.input_251==999) or iparam.dtnowstr[5:]>"04-10":
        but_disabled='disabled style="BACKGROUND-COLOR: grey; '
    else:
        but_disabled=' style="BACKGROUND-COLOR: green;'
    ret += u'<button title="only active from January 1st until April 10th EOB" %s height: 50px;color: white; font-weight: bold" onclick="getByJForm(\'BAV_Yearly.html\',null,\'%s\',true);">Edit your yearly conversion</button><br><br>' %  (but_disabled,object2bas64(extra))


    iparam.otype = "standorthis"
    ret += "<hr/><h7></h7>\n<h2>Standorthistorie<h2>" + handleHTML(iparam, getOCACHE()["standort_qnr"].get(qnum), 0, 0,attribs=["q_nummer","gueltig_ab","strasse","hausnr","adresszusatz","plz","wohnort","gebaeudekuerzel","anzahl_einfache_km"],otype="standorthis")    
    stndort = getStandortByDate(qnum, iparam.dtnowstr)
    extra = CONT()
    extra.name = getMAName(uma)
    extra.__textreplace = CONT()
    extra.__textreplace.PHR_name = getMAName(uma)
    extra.q_nummer = qnum
    extra.tstaette = stndort.gebaeudekuerzel if stndort else ""
    ret += u'<button style="BACKGROUND-COLOR: green; height: 50px;color: white; font-weight: bold" onclick="getByJForm(\'Change_of_residence.html\',null,\'%s\',true);">Enter new address</button><br><br>' %  object2bas64(extra) #base64.b64encode(dumps(extra))
    iparam.otype = "vertragshis"
    ret += "<hr/><h7></h7>\n<h2>Vertragshistorie<h2>" + handleHTML(iparam, getOCACHE()["vertrag_qnr"].get(qnum), 0, 0,attribs=["q_nummer","gueltig_ab","vertragsart","fuehrungskraft","konzernleihe","team","abteilung","zielabteilung"])    
    iparam.otype = "abwesenheit"
    vals = getOCACHE()["abwesenheit_qnr"].get(qnum)
    ret += "<hr/><h7></h7>\n<h2>Abwesenheitshistorie<h2>" + handleHTML(iparam, vals, 0, attribs=["q_nummer","start","ende","geplanter_geburtstag_muttersch","tatsaechlicher_geburtstag_mutt","art","name_des_kindes"])   
    iparam.wbname = "dienstfahrzeuge"
    iparam.otype = "dienstfahrzeuge"
    vals = getOCACHE()["dienstfahrzeuge_qnr"].get(qnum)
    ret += "<hr/><h7></h7>\n<h2>Dienstfahrzeuge<h2>" + handleHTML(iparam, vals, 0,attribs=["q_nummer","kennz","typ","full_service_rate","finanzrate","gesamt","brutto_preis","uebernahme","rueckgabe","leasing_beginn","leasing_ende","aenderungsgrund","rate_1_vom_blp","dfzg_typ","vin","leasing_nr"])
    if 1:  # cb1:
        cb = CONT()
        cb.oid = qnum  # cb1.oid
        cb.q_nummer = qnum  # cb1.oid
        from shared_code.Services import getCompbike
        reload(getCompbike)
        vals = list(getCompbike.getCompbike(iparam, [cb]).values())
        ret += "<hr/><h7></h7>\n<h2>Company Bike Entitlement<h2>" + handleHTML(iparam, vals, 0, 1, ["q_nummer", "name", "vertragsart", "eintritt", "austritt", "berechtigung"])
    
    return ret

def handleHTML(iparam, vals, transpone=0, header=1, attribs=None,otype=None,attribs_trans=None,extra_styles=None):
    if attribs_trans:
        attribs=attribs_trans.keys()
    if not otype and iparam.otype:otype=iparam.otype
    if not vals: return ""
    if vals == "GetPersonalInfo":
        return handleGenHTML(iparam)
    if attribs == None:
        attribs = getAttributes(iparam)
    
    html = u""
    if header:
        html += """<head>  <meta charset="cp1252"> <title>Personal Info</title> 
    <style> @media print { h7 {page-break-after:always;} } </style>  
<style>
    * {
      font-family: sans-serif;
      font-size:97%;
    }

table, th, td {
    border: 2px solid black;
    border-collapse: collapse;
    white-space:pre-wrap; word-break:break-word;
}
th {
  word-break:break-all;
  white-space:normal;
  word-wrap:break-word;
  BACKGROUND-COLOR: #bbf0b6;
}

hr {
    display: block;
    height: 1px;
    border: 0;
    border-top: 2px solid;
    margin: 1em 0;
    padding: 0;
}</style>    </head> """

    headmap={}
    if otype:
        smodule=XEDB.DYNMODS.get(otype)
        if smodule:
            if hasattr(smodule,"UIGRID"):
                gridinfo=smodule.UIGRID
            if gridinfo:
                for f in gridinfo["fields"]:
                    headmap[f["name"]]=f["text"]

    html += '<table>\n<tr>'
    printed = []
    if transpone:
        html += '<tr>\n'
        for a in attribs:
            if headmap:
                mappeda=headmap.get(a,a)
            else:mappeda=a
            html += '<td style="BACKGROUND-COLOR: #bbf0b6;"><b>%s</b></td>' % mappeda
            for r in vals:
                val = ""
                if hasattr(r, a):
                    val = getattr(r, a)
                if val == "#NV":val = ""
                if val and val.endswith(".0"):val = val[:-2]
                html += "<td>%s</td>" % val
            html += '<tr>\n'
    else:
        if attribs:
            for a in attribs:
                if headmap:
                    mappeda=headmap.get(a,a)
                    mappeda=mappeda.replace("  "," ")
                    mappeda=mappeda.replace(" -","<br>")
                    mappeda=mappeda.replace("- ","<br>")
                    mappeda=mappeda.replace("-","<br>")
                    mappeda=mappeda.replace(" ","<br>")
                else:mappeda=a
                extra_style=""
                if extra_styles and extra_styles.get(mappeda):
                    extra_style=extra_styles.get(mappeda)
                if attribs_trans and attribs_trans.get(mappeda):
                    mappeda=attribs_trans.get(mappeda)
                html += '<th %s><b >%s</b></th>' % (extra_style,mappeda)
            html += '<tr>\n'
            if "q_nummer" in attribs and "gueltig_ab" in attribs:
                vals = sorted(sorted(vals, key=lambda x: x.gueltig_ab, reverse=True), key=lambda x: x.q_nummer, reverse=False)
            elif "q_nummer" in attribs and "year" in attribs:
                vals = sorted(sorted(vals, key=lambda x: x.year, reverse=True), key=lambda x: x.q_nummer, reverse=False)
            elif "q_nummer" in attribs:
                vals.sort(key=lambda x: x.q_nummer, reverse=False)
            for r in vals:
                for a in r.__dict__:
                    if a not in printed:
                        printed.append(a)
                html += "<tr>"
                for a in attribs:
                    val = ""
                    if hasattr(r, a):
                        val = getattr(r, a)
                    if val == "#NV":val = ""
                    val=str(val)
                    if val and val.endswith(".0"):val = val[:-2]
                    html += "<td>%s</td>" % val
                html += '<tr>\n'
    html += "</table>"
    return html

def MA_DICT():
    return getOCACHE()["mitarbeiter_qnr"]



def refreshDicts():
    getOCACHE()["abtlg_kuerzel"]=dictByAttrList("abteilung","kuerzel")
    getOCACHE()["team_kuerzel"]=dictByAttrList("team","kuerzel")
    getOCACHE()["hauptabteilung_kuerzel"]=dictByAttrList("hauptabteilung","kuerzel")
    getOCACHE()["vertrag_qnr"]=dictByPNumList("vertragshis")
    getOCACHE()["mitarbeiter_qnr"]=dictByPNum("mitarbeiter")
    getOCACHE()["nomiko_qnr"]=dictByPNum("nomiko")
    getOCACHE()["planzak_qnr"]=dictByPNum("planzak")
    getOCACHE()["planzak_qxnr"]=dictByAttr("planzak","qx_nummer")    
    getOCACHE()["gehalt_qnr"]=dictByPNumList("gehalt")
    getOCACHE()["abwesenheit_qnr"]=dictByPNumList("abwesenheit")
    getOCACHE()["absence_qnr"]=dictByPNumList("absence")
    getOCACHE()["standort_qnr"]=dictByPNumList("standorthis")
    getOCACHE()["dienstfahrzeuge_qnr"]=dictByPNumList("dienstfahrzeuge")

def handle(iparam,internal=0):
    while not XEDB.CACHE_LOADED:
        time.sleep(2)
    addIparamVals(iparam)    
    refreshCacheAll(iparam)

    if 0 or not XEDB.DICTREFRESH:
        refreshDicts()
        XEDB.DICTREFRESH=time.localtime()

    if XEDB.ENVIRONMENT in ["TEST"]:
        cwd=os. getcwd()
        if "azurefunction" not in cwd:
            os.chdir("azurefunction")

    
    # OLD Version: get Login / Changeuser Stuff from DB
    DBROLES=[]
    linf=getLinf(iparam)
    if linf and linf.roles:
        DBROLES=linf.roles

    ret=CONT()

    if (XEDB.AZURE and not internal and iparam.otype not in ["open_charge","ping","jotform",'sap_absence_upload']) or (not internal and not linf and iparam.otype not in ["open_charge","ping",'sap_absence_upload']):
        if XEDB.DOMAIN!="carit" or XEDB.AZURE:
            import shared_code.oidConnect
            if XEDB.HOSTNAME in ["PITSNUC","PITSLG"]:
                reload(shared_code.oidConnect)
            linf=shared_code.oidConnect.checkAuth(iparam)
        else:
            import requests
            headers = {'content-type':'application/json','Cookie': 'phrAccessToken=%s'%iparam.xsecookie}
            data='{"user":"WappLogin","domain":"'+XEDB.DOMAIN+'","method":"GetValues","otype":"authwapp","client":"powerbi"}'
            response=requests.post("https://127.0.0.1",headers=headers,data=data,verify=False)
            if 0:
                authok=0
                try:
                    robj=loads(response.text)
                    if robj != []: authok=1
                except:pass
            if 1 or authok:
                refreshCacheAll(iparam)
                linf=getLinf(iparam)
        if linf and linf.RetCode== "Login failed" and iparam.otype=="jotform":
            linf=CONT()
            linf.user=pnum=iparam.user="anonymous"
            linf.roles=[]
        elif linf and linf.RetCode== "Login failed":
            return linf
        if (not linf and not internal):
            ret=CONT()
            ret.RetCode="Authentication failed"
            return ret

        if not linf.user:
            linf.user=linf.amuser.sub if XEDB.DOMAIN=="carit" else linf.amuser.v_num
        if linf.user!="anonymous":
            pnum=linf.amuser.q_nummer if XEDB.DOMAIN=="carit" else linf.amuser.v_num
            iparam.user = pnum
        if not linf.roles or linf.roles==["changeuser"]:
            loadRoles()
            was_changeuser="changeuser" in linf.roles
            if 0 and DBROLES:
                linf.roles=DBROLES
            else:
                linf.roles=XEDB.USERROLES.get(pnum,[])
            improles=implicitRoles(pnum,iparam.dtnowstr)
            linf.roles.extend(improles) 
            if was_changeuser and "changeuser" not in linf.roles:
                linf.roles.append("changeuser")
        extauth = XPY_Auth(iparam, linf)
        #logging.info("WappHandler line 3375: iparam.extauth is %s" % (extauth))
        logging.info("%s has roles %s"%(linf.user,dumps(linf.roles)))
        if extauth and iparam.method=="Login":
            ret.RetCode = "ok"
            ret.XSECOOKIE=linf.xsecookie
            return ret
        if iparam.method=="DeleteObject" and "hreditor" not in linf.roles:
            totype=getOtypeByOID(iparam.oid)
            if totype in ["absence","pofofb"]:
                ao=getObject(totype, iparam.oid)
                extauth=isMyObject(ao,linf,totype)
            if totype in ['dienstfahrzeuge'] and 'servicesFuhrpark' in linf.roles:
                extauth=1
            if totype in ['trainings']:
                extauth= 'servicesTrainings' in linf.roles

        #logging.info("WappHandler line 3389: iparam.extauth is %s" % (extauth))                             
        iparam.extauth = extauth
        if not extauth and iparam.method=="GetValues" and iparam.otype=="notify" and iparam.user=="notifier":
            extauth=1
        if not extauth and iparam.method=="GetValues" and iparam.otype=="jotform" and iparam.user=="anonymous":
            extauth=1
        if not extauth and iparam.method in ['GetUserInfo','GetAppList',"GetMAList","XLSDownload"]:
            extauth=1
        if not extauth:
            hint="###role not sufficient for method %s - %s - user: %s on domain: %s" % (iparam.method, iparam.otype, iparam.user,XEDB.DOMAIN)
            logging.info(hint)
            ret.RetCode = hint
            #ret.RetCode = "You have been logged out (probably by an app restart). Please log in again on the start page."
            return ret


    ret=handleFieldHistory(iparam)
    if ret !=None:return ret
    ret=handleDeleteObject(iparam)
    if ret:return ret
    olist,ret=handleSetValues(iparam)
    if ret and ret.RetCode!="ok":return ret
    if ret and ret.GridReload and iparam.otype in ["meter_reading"]:
        return ret
    RetCodeStore=None
    if ret:RetCodeStore=ret.RetCodeStore
    ret=handleSpecialMethods(iparam,linf)
    if ret:return ret
    jotform=XEDB.JOTFORMS.get(iparam.otype)
    if jotform:
        ret1=CONT()
        ret1.html='    <iframe       id="JotFormIFrame"       title=""       onload="window.parent.scrollTo(0,0)"       allowtransparency="true"       allowfullscreen="true"       allow="geolocation; microphone; camera"       src="forms/' + jotform + '"       frameborder="0"       style="       min-width: 100%;       min-height:100%;       border:none;"       scrolling="yes"     zoom: 100%"> '
	
    else:
        func=getFunc(iparam)
    if iparam.method in ["GetValues","SetValues"] and not iparam.download:
        org_otype=iparam.otype
        if not jotform:
            ret1=func(iparam,olist)
        ret=CONT()
        ret.retd={}
        ret.retd[iparam.otype]=CONT()
        if isinstance(ret1,CONT) and (org_otype in ["charging_upload","GetPersonalInfo",'sap_absence_upload'] or iparam.download):
            if ret1 and ret1.filename:#org_otype in ["hr_pro","'fporeport'"]:
                handleReports(iparam,ret1) 
            return ret1
        if org_otype in ['GetFINPersons']:
            return ret1
        if isCont(ret1) and ret1.html_direct:
            return ret1
        if 0 and isCont(ret1) and ret1.RetCode and ret1.RetCode!="ok":
            return ret1
        if isinstance(ret1,CONT) and (ret1.html or ret1.vega or ret1.RetCode or  ret1.ReloadPage):
            ret.retd[iparam.otype]=ret1
            if 0 and ret1.vega:
                ret1.html=1
            if 0 and ret1.vega:
                ret.gridinfos={iparam.otype:{"html":1,"fields":[]}}
                return ret
            gridinfo=None
            if not jotform:
                if hasattr(iparam,"unreleased_otype"):
                    smodule=XEDB.DYNMODS[iparam.unreleased_otype]
                else:
                    smodule=XEDB.DYNMODS[iparam.otype]
                if hasattr(smodule,"getUIGRID"):
                    gridinfo=smodule.getUIGRID(iparam)
                else:
                    if hasattr(smodule,"UIGRID"):
                        gridinfo=smodule.UIGRID
            if isCont(ret1) and((ret1.RetCode and ret1.RetCode!="ok") or ret1.ReloadPage):
                if gridinfo:
                    ret1.gridinfos={org_otype:gridinfo}
                return ret1
            if not ret1.vega:
                if gridinfo and "inputfields" in gridinfo:
                    ret.gridinfos={iparam.otype:{"html":1,"fields":[],"inputfields":gridinfo["inputfields"]}}
                else:
                    ret.gridinfos={iparam.otype:{"html":1,"fields":[]}}
            if ret1.RetCode:
                ret.RetCode=ret1.RetCode
        else:
            ret.retd[iparam.otype].odict=ret1
    else:
        ret=func(iparam,olist)
    if not ret.gridinfos and iparam.method=="GetValues" and hasattr(iparam,"gridinfo"):
        handleReports(iparam,ret) 
    if ret.gridinfos:
        fillGridCombos(ret)
    if hasattr(iparam,"cfstatus") and hasattr(ret,"retd") and iparam.otype in ret.retd: 
        ret.retd[iparam.otype].cfstatus=iparam.cfstatus
    if iparam.client=="powerbi":
        ret=list(list(ret.retd.values())[0].odict.values())
    if iparam.client=="singleobject":
        ret=list(ret.retd.values())[0]

    if  ret and RetCodeStore:
        ret.RetCodeStore=RetCodeStore

    ### if a file dump is needed:
    if 0:
        fwrite=open("stream.json", "w")
        fwrite.write(dumps(ret))
        fwrite.close()

    if 1 and hasattr(iparam,"org_otype") and iparam.org_otype !=iparam.otype and iparam.method!='XLSDownload':
        ret.retd[iparam.org_otype]=ret.retd[iparam.otype]
        del ret.retd[iparam.otype]
        iparam.otype=iparam.org_otype

    if ret.RetCode and ret.retd:
        del ret.retd

    return ret

PRTABLE={"PL6":30000,"PL5":14700,
"PL4":10300,
"PL3":6100,
"PL2":5400,
"PL1":5400,
"SL4":7100,
"SL3":5400,
"SL2":3900,
}
def copyUIGRIDAttribs(r,o,uigrid):
    for uif in uigrid["fields"]:
        setFValue(r, o, uif["name"])

def copyAttribs2(m,g,attl=None,repl_nl=0,nvblank=0):
    if not g:
        return
    for a in g.__dict__:
        if attl != None:
            if a not in attl:continue
        val=getattr(g, a)
        if nvblank and val=="#NV":continue
        if repl_nl:
            val=val.replace("\n","<br>")
        setattr(m, a, val)

def getHolidaysGermany(thisyear, bundesland):
    if XEDB.HOLDAY_GER==None:
        XEDB.HOLDAY_GER={}
    ret=XEDB.HOLDAY_GER.get((thisyear, bundesland))
    if not ret:
        ret=dict(holidays.Germany(years=int(thisyear), prov=bundesland))
        XEDB.HOLDAY_GER[(thisyear, bundesland)]=ret
    return ret

def appendUnique(ll,el):
    if el not in ll:
        ll.append(el)

def getHolidaysGermanyShort(thisyear, bundesland):
    if XEDB.HOLDAY_GER_SHORT==None:
        XEDB.HOLDAY_GER_SHORT={}
    ret=XEDB.HOLDAY_GER_SHORT.get((thisyear, bundesland))
    if not ret:
        hlong=getHolidaysGermany(thisyear, bundesland)
        ret={}
        for d,f in hlong.items():
            ret[date2String(d)[5:]]=f
        XEDB.HOLDAY_GER_SHORT[(thisyear, bundesland)]=ret
    return ret

def getOrAdd(dd,k,v):
    ret=dd.get(k)
    if ret:return ret
    dd[k]=v
    return dd[k]

def listAsMem(ll):
    ret=CONT()
    i=0
    for m in ll:
        setattr(ret,str(i),m)
        i+=1
    return ret

def getOCACHE():
    """TODO: Add documentation.

    Returns: XEDB.OCACHE
    """
    tid=threading.currentThread().ident    
    if tid in XEDB.TCACHES:
        return XEDB.TCACHES[tid]
    return XEDB.OCACHE

def safe_delattr(obj, attrname):
    if hasattr(obj, attrname):
        delattr(obj, attrname)

def setTCACHE():
    tid=threading.currentThread().ident    
    XEDB.TCACHES[tid]={}
    
def delTCACHE():
    dummy=XEDB
    tid=threading.currentThread().ident
    if tid in XEDB.TCACHES:
        del XEDB.TCACHES[tid]


def maAktiv(iparam,ma,r=None,adatumstr=None):
    if getS(ma,"inaktiv"):
        if r:
            r.aktiv=""
        return False
    if not iparam:
        iparam=CONT()
        if adatumstr:
            iparam.adatumstr=adatumstr
        else:
            iparam.adatumstr=datetime.date.today().strftime("%Y-%m-%d")
    ad=getS(ma,"austrittsdatum")
    if not ad:
        ad=getS(ma,"ende_anue_datum")
    ed=getS(ma,"eintrittsdatum")
    if not ed:
        ed=getS(ma,"beginn_anue_datum")
    if not ad: ad="9999-12-31"
    if not ed: ed="9999-12-31"
    ret=False
    if iparam.adatumstr<=ad and iparam.adatumstr>=ed:
        if r:
            r.aktiv="X"
        ret=True
    else:
        if r:
            r.aktiv=""
    return ret



def getUsedAttribs(rows,vattribs):
    for r in rows:
        for a in r.__dict__:
            if a =="rows":
                getUsedAttribs(r.rows.values(),vattribs)
            else:
                if a not in vattribs:
                    vattribs.append(a)

def handleReports(iparam,ret):
    if 1: #iparam.otype in ["devtest","history","mitarbeiter_all","urlaub","bem"]: #"urlaub",
        if hasattr(iparam,"gridfiltertype") and iparam.gridfiltertype=="Filter2": 
            allreadonly=0
        else: allreadonly=0
        attribs=params=None
        gridinfo=0#XEDB.UIGRIDS.get(iparam.otype)
        if not gridinfo:
            try:
                if hasattr(iparam,"unreleased_otype"):
                    smodule=XEDB.DYNMODS[iparam.unreleased_otype]
                else:
                    smodule=XEDB.DYNMODS[iparam.otype]
                if hasattr(smodule,"getUIGRID"):
                    gridinfo=smodule.getUIGRID(iparam)
                else:
                    if hasattr(smodule,"UIGRID"):
                        gridinfo=smodule.UIGRID
                if hasattr(iparam,"download") and not ret.filename:return # iparam.otype not in ["hr_pro"]: return
            except:
                print(traceback.format_exc())
        if gridinfo:
            fields=gridinfo.get("fields")
        else:
            fields=None
        if fields:
            linf=getLinf(iparam)
            if 0 and iparam.otype=="talent" and "hrlead" in linf.roles:
                fields.append({"text":u"Sel", "name": "sel", "type": "ed", "width":  6})
            attribs=[]
            fields2=[]
            for f in fields:
                if linf and iparam.otype=="talent" and f["name"]=="bewqnum" and "hrlead" not in linf.roles:
                    continue
                if linf and iparam.otype=="talent" and f["name"]=="status" and "hrlead" in linf.roles:
                    f["type"]="coro"
                fields2.append(f)
                attribs.append(f["name"])
            fields=fields2
        vattribs=[]
        if hasattr(iparam,"org_otype"):
            ret.retd[iparam.org_otype]=ret.retd[iparam.otype]
            del ret.retd[iparam.otype]
            iparam.otype=iparam.org_otype
        if not attribs:
            try:
                odict=ret.retd.values()[0].odict
                getUsedAttribs(odict.values(),vattribs)
            except:pass
        if attribs:
            if fields:
                fieldsd={}
                for f in fields:
                    fieldsd[f["name"]]=f
            ret.gridinfos={}
            ret.gridinfos[iparam.otype]=CONT()
            ret.gridinfos[iparam.otype].fields=[] #{}
            if gridinfo:
                for k in gridinfo.keys():
                    if k not in ["fields"]:
                        setattr(ret.gridinfos[iparam.otype],k,gridinfo.get(k))
            i=0
            for a in attribs:
                i+=1
                f=CONT()
                f.text=a
                f.name=a
                f.type="ro"
                fwidth=10
                if fields:
                    finfo=fieldsd.get(a)
                    if finfo and "type" in finfo:
                        if vattribs and finfo["type"]=="ro" and a not in vattribs:continue
                        if finfo["type"] != "hidden" and allreadonly:
                            f.type="ro"
                        else:
                            f.type=finfo["type"]
                    if finfo and "text" in finfo:
                        f.text=finfo["text"]
                    if finfo and "values" in finfo:
                        f.values=finfo["values"]
                    if finfo and "width" in finfo:
                        fwidth=finfo["width"]
                    if finfo and "sort" in finfo:
                        f.sort=finfo["sort"]
                    if finfo and "color" in finfo:
                        f.color=finfo["color"]
                    if finfo and "exttype" in finfo:
                        f.exttype=finfo["exttype"]
                    if finfo and "byfunction" in finfo:
                        f.byfunction=finfo["byfunction"]
                    if finfo and "editorwidth" in finfo:
                        f.editorwidth=finfo["editorwidth"]
                f.width=fwidth
                ret.gridinfos[iparam.otype].fields.append(f)
            if params:
                ret.gridinfos[iparam.otype].inputfields=gridinfo["inputfields"]

def getGFuehrer(cdate):
    for gf in getObjects("gmbh"):
        if gf.gueltig_bis and  cdate>gf.gueltig_bis:continue
        if gf .gueltig_ab and cdate>=gf.gueltig_ab:
            return gf

def getMAName(m,short=0):
    if not m: return "#NV"
    if isinstance(m,str):
        qnr=m
        for motype in ["mitarbeiter","planzak","nomiko"]:
            m=objFromCache("%s_qnr"%motype,qnr)
            if m:break
        if not m:
            pzakd=getOCACHE().get("planzak_qxnr")
            if not pzakd:
                pzakd=getOCACHE()["planzak_qxnr"]=dictByAttr("planzak","qx_nummer")
            m=pzakd.get(qnr)

    if not short:
        return "%s, %s"%(getS(m,"nachname"),getS(m,"vorname"))        
    vorname=getS(m,"vorname")
    vorname=vorname[0] if vorname else vorname
    return "%s %s."%(getS(m,"nachname"),vorname)        

def convFloatFormat(f):
    if not f:return f
    dotfnd=f.find(".")+1
    commafnd=f.find(",")+1
    if dotfnd and commafnd:
        f=f.replace(".","")
    if commafnd:
        f=f.replace(",",".")
    return f

def date2String(d1):
    if not d1:return None
    if isinstance(d1, str): return d1
    return d1.strftime(DFMT)

class HISTORY:
    def __init__(self,otype,attr="q_nummer") -> None:
        self.ddict=dictByAttrList(otype,attr)
        self.attr=attr
        for gl in self.ddict.values():
            gl.sort(key=lambda x: setNoneAs(x.gueltig_ab,"0000-00-00"), reverse=True)
    def get(self,attr,dstring):
        if repr(type(dstring)) == "<class 'datetime.date'>":dstring = dstring.strftime(DFMT)
        if repr(type(dstring)) == "<class 'datetime.datetime'>":dstring = dstring.strftime(DFMT)
        gl=self.ddict.get(attr,[])
        for g in gl:
            if not g.gueltig_ab:continue
            try:
                if dstring >= g.gueltig_ab:
                    return g
            except:
                raise


def getCityByDate(qnr, dstring):
    try:
        o = getStandortByDate(qnr, dstring)
        g = objFromCache("gebaeude", o.gebaeudekuerzel)
        return g.standort
    except:
        return "#NV"

def string2Date(d1):
    if repr(type(d1))=="<class 'datetime.date'>":return d1
    d1=d1.strip()
    sd=d1.split("-")
    return datetime.date(int(sd[0]),int(sd[1]),int(sd[2]))

def addDayDiff2String(d,days):
    d1=string2Date(d)
    d2=d1+timedelta(days=days)
    return d2.strftime(DFMT)

def string2DateTime(d1):
    if repr(type(d1))=="<class 'datetime.date'>":return d1
    d1=d1.strip()
    sd=d1.split(" ")
    sd1=sd[0].split("-")
    sd2=sd[1].split(":")
    ret=datetime.datetime(int(sd1[0]),int(sd1[1]),int(sd1[2]),int(sd2[0]),int(sd2[1]),int(sd2[2]))
    return ret

def getN(r,f,n=0):
    try:
        rv=getattr(r, f)
        if isCont(rv):rv=rv.value
        if isinstance(rv, str):return float(convFloatFormat(getattr(r, f)))
        else: return float(rv)
    except:
        try:
            return float(n)
        except: return n

def getD(r,f=None,n=None):
    '''Magically generates a date from a string or an object.
    Sometimes works, sometimes fails. Used in several critical places
    all over the system.'''
    try:
        if f:
            d1=getattr(r, f)
        else:
            d1=r
        sd=d1.split("-")
        return datetime.date(int(sd[0]),int(sd[1]),int(sd[2]))
    except:
        return n

def getS(r,f,n="",empty_is_none=False):
    """TODO: Add documentation.

    Returns: n (str)
    """
    try:
        ret=getattr(r, f)
        if ret=="" and not empty_is_none:
            return ret
        if not ret:
            if n:
                return n
            if not empty_is_none:
                return n
        return ret
    except:
        return n

def fillCombo(ginfos,fname,sotype,sofield):
    for f in ginfos.fields:
        if f.name==fname:
            vals=[]
            for o in getObjects(sotype):
                if not hasattr(o,sofield):continue
                v=getattr(o,sofield)
                if v not in vals:
                    vals.append(v)
            f.values=sorted(vals)
            if fname in ["aktuelle_einstufung"]:
                f.values.append("")

def fillGridCombos(ret):
    for otype,ginfos in ret.gridinfos.items():
        if otype in ["planzak","nomiko"]:
            fillCombo(ginfos,"abteilung","abteilung","kuerzel")
            fillCombo(ginfos,"team","team","kuerzel")
        if otype=="vertragshis":
            fillCombo(ginfos,"team","team","kuerzel")
            fillCombo(ginfos,"abteilung","abteilung","kuerzel")
        if otype.startswith("bewerber"):
            fillCombo(ginfos,"teamkuerzel","team","kuerzel")
            fillCombo(ginfos,"abteilung","abteilung","kuerzel")
        if otype=="gehalt":
            fillCombo(ginfos,"aktuelle_einstufung","kompetenzstufe","k_stufe")
        if otype=="standorthis":
            fillCombo(ginfos,"gebaeudekuerzel","gebaeude","kuerzel")
        if otype=="team":
            fillCombo(ginfos,"abteilung","abteilung","kuerzel")
